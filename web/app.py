#!/usr/bin/env python3
"""
Web Dashboard — Bollinger Band Squeeze Strategy Analyser
Flask backend that exposes the same analysis as the terminal app via a
beautiful Bootstrap 5 + Chart.js web interface.

Run:  python web/app.py
Then open:  http://127.0.0.1:5000
"""

import sys, os, json, math, threading
from dataclasses import asdict
from datetime import date, timedelta
import pandas as pd
import numpy as np
import time as _time

# Load .env file if present (for auth config, SMTP, Google OAuth, etc.)
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))
except ImportError:
    pass

# ── Ensure project root is on path ──────────────────────────────
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from flask import Flask, render_template, jsonify, request, Response, g
from flask.json.provider import DefaultJSONProvider


class _NumpySafeJSONProvider(DefaultJSONProvider):
    """Extend Flask's default JSON provider to handle numpy types."""

    @staticmethod
    def default(o):
        if isinstance(o, (np.bool_,)):
            return bool(o)
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, (np.floating,)):
            v = float(o)
            if math.isnan(v) or math.isinf(v):
                return None
            return v
        if isinstance(o, np.ndarray):
            return o.tolist()
        return DefaultJSONProvider.default(o)

from bb_squeeze.data_loader import (
    normalise_ticker, load_stock_data, get_all_tickers_from_csv, get_data_freshness,
)
from bb_squeeze.indicators import compute_all_indicators
from bb_squeeze.signals import analyze_signals
from bb_squeeze.fundamentals import fetch_fundamentals
from bb_squeeze.strategies import run_all_strategies, strategy_result_to_dict
from bb_squeeze.quant_strategy import run_quant_analysis
from bb_squeeze.config import CSV_DIR
from hybrid_pa_engine import run_triple_analysis
from price_action.engine import run_price_action_analysis, pa_result_to_dict
from bb_squeeze.trade_db import init_db as _init_trade_db, add_trade, get_all_trades, get_trade, delete_trade, update_trade, user_has_trades, delete_trades_by_position
from bb_squeeze.trade_calculator import calculate_trade, calculate_fy_summary
from bb_squeeze.portfolio_db import (
    init_portfolio_db as _init_portfolio_db,
    add_position, get_all_positions, get_open_positions, get_closed_positions,
    get_position, update_position, close_position, reopen_position, delete_position,
    user_has_positions,
)
from bb_squeeze.portfolio_analyzer import analyze_position, analyze_all_open_positions
from web.ta_routes import ta_bp
from web.hybrid_routes import hybrid_bp
from web.top_picks_routes import top_picks_bp
from web.pa_routes import pa_bp
from web.triple_routes import triple_bp
from web.vince_routes import vince_bp
from web.mental_game_routes import mental_game_bp
from web.rentech_routes import rentech_bp
from web.auth_routes import auth_bp
from web.sentiment_routes import sentiment_bp
from mental_game.db import init_mental_game_db as _init_mental_game_db
from auth.db import init_auth_db as _init_auth_db
from auth.middleware import init_auth_middleware

app = Flask(__name__)
app.json_provider_class = _NumpySafeJSONProvider
app.json = _NumpySafeJSONProvider(app)
app.register_blueprint(ta_bp)
app.register_blueprint(hybrid_bp)
app.register_blueprint(top_picks_bp)
app.register_blueprint(pa_bp)
app.register_blueprint(triple_bp)
app.register_blueprint(vince_bp)
app.register_blueprint(mental_game_bp)
app.register_blueprint(rentech_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(sentiment_bp)

# Global Macro / Inter-Market Sentiment (separate, isolated module)
from web.global_sentiment_routes import global_sentiment_bp
app.register_blueprint(global_sentiment_bp)

# Bootswatch theming
from web.theme import theme_bp, inject_theme
app.register_blueprint(theme_bp)
app.context_processor(inject_theme)

# Ensure DBs exist
_init_trade_db()
_init_portfolio_db()
_init_mental_game_db()
_init_auth_db()

# ── Auth middleware (must be AFTER all blueprints are registered) ──
init_auth_middleware(app)


# ── Inject user context into all templates ──────────────────────
@app.context_processor
def inject_user_context():
    """Make user info available in every template."""
    user = getattr(g, "user", None)
    if not user:
        return {"current_user": None, "is_admin": False, "user_initials": "", "avatar_color": "#58a6ff"}
    fn = (user.get("first_name") or "")[:1].upper()
    ln = (user.get("last_name") or "")[:1].upper()
    initials = (fn + ln) or (user.get("username") or "U")[:2].upper()
    # Deterministic color from username hash
    import hashlib
    h = int(hashlib.md5((user.get("username") or "").encode()).hexdigest()[:6], 16)
    colors = ["#f44336","#e91e63","#9c27b0","#673ab7","#3f51b5","#2196f3","#00bcd4","#009688","#4caf50","#ff9800","#ff5722","#795548"]
    color = colors[h % len(colors)]
    return {
        "current_user": user,
        "is_admin": user.get("is_admin", False),
        "user_initials": initials,
        "avatar_color": color,
    }


def _uid():
    """Get current user_id from request context."""
    u = getattr(g, "user", None)
    return u.get("user_id") if u else None


def _is_admin():
    """Check if current user is admin."""
    u = getattr(g, "user", None)
    return bool(u and u.get("is_admin"))


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────

def _safe(v, decimals=2):
    """Return a JSON-safe value (None / rounded float / string)."""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return round(v, decimals)
    return v


def _fmt_cr(val):
    """Format raw ₹ value as Crores string."""
    if val is None:
        return "N/A"
    cr = val / 1e7
    if cr >= 1_00_000:
        return f"₹{cr/1_00_000:.2f}L Cr"
    if cr >= 1_000:
        return f"₹{cr/1_000:.1f}K Cr"
    return f"₹{cr:.2f} Cr"


def _build_chart_data(df):
    """Build OHLCV + indicator series for charts."""
    # Use last 250 trading days (~1 year)
    df = df.tail(250).copy()
    dates = [d.strftime("%Y-%m-%d") for d in df.index]

    return {
        "dates":    dates,
        "open":     [round(float(v), 2) for v in df["Open"]],
        "high":     [round(float(v), 2) for v in df["High"]],
        "low":      [round(float(v), 2) for v in df["Low"]],
        "close":    [round(float(v), 2) for v in df["Close"]],
        "volume":   [int(v) for v in df["Volume"]],
        "bb_upper": [round(float(v), 2) if not math.isnan(v) else None for v in df["BB_Upper"]],
        "bb_mid":   [round(float(v), 2) if not math.isnan(v) else None for v in df["BB_Mid"]],
        "bb_lower": [round(float(v), 2) if not math.isnan(v) else None for v in df["BB_Lower"]],
        "sar":      [round(float(v), 2) if not math.isnan(v) else None for v in df["SAR"]],
        "sar_bull": [bool(v) for v in df["SAR_Bull"]],
        "bbw":      [round(float(v), 6) if not math.isnan(v) else None for v in df["BBW"]],
        "pctb":     [round(float(v), 4) if not math.isnan(v) else None for v in df["Percent_B"]],
        "cmf":      [round(float(v), 4) if not math.isnan(v) else None for v in df["CMF"]],
        "mfi":      [round(float(v), 2) if not math.isnan(v) else None for v in df["MFI"]],
        "vol_sma":  [round(float(v), 0) if not math.isnan(v) else None for v in df["Vol_SMA50"]],
        "squeeze":  [bool(v) for v in df["Squeeze_ON"]],
    }


def _signal_dict(sig):
    """Convert SignalResult to a clean JSON-serialisable dict."""
    return {
        "ticker":           sig.ticker,
        "phase":            sig.phase,
        "current_price":    _safe(sig.current_price),
        "bb_upper":         _safe(sig.bb_upper),
        "bb_mid":           _safe(sig.bb_mid),
        "bb_lower":         _safe(sig.bb_lower),
        "bbw":              _safe(sig.bbw, 6),
        "bbw_6m_min":       _safe(sig.bbw_6m_min, 6),
        "percent_b":        _safe(sig.percent_b, 4),
        "sar":              _safe(sig.sar),
        "sar_bull":         sig.sar_bull,
        "volume":           int(sig.volume),
        "vol_sma50":        int(sig.vol_sma50) if sig.vol_sma50 else 0,
        "cmf":              _safe(sig.cmf, 4),
        "mfi":              _safe(sig.mfi, 2),
        "cond1":            sig.cond1_squeeze_on,
        "cond2":            sig.cond2_price_above,
        "cond3":            sig.cond3_volume_ok,
        "cond4":            sig.cond4_cmf_positive,
        "cond5":            sig.cond5_mfi_above_50,
        "buy_signal":       sig.buy_signal,
        "sell_signal":      sig.sell_signal,
        "hold_signal":      sig.hold_signal,
        "wait_signal":      sig.wait_signal,
        "head_fake":        sig.head_fake,
        "confidence":       sig.confidence,
        "direction_lean":   sig.direction_lean,
        "squeeze_days":     sig.squeeze_days,
        "stop_loss":        _safe(sig.stop_loss),
        "summary":          sig.summary,
        "action_message":   sig.action_message,
        "exit_sar_flip":    sig.exit_sar_flip,
        "exit_lower_band":  sig.exit_lower_band_tag,
        "exit_double_neg":  sig.exit_double_neg,
        # ── New Book Indicators (Ch.15, 18, 21) ──
        "ii_pct":           _safe(sig.ii_pct, 4),
        "ad_pct":           _safe(sig.ad_pct, 4),
        "vwmacd_hist":      _safe(sig.vwmacd_hist, 4),
        "expansion_up":     sig.expansion_up,
        "expansion_down":   sig.expansion_down,
        "expansion_end":    sig.expansion_end,
        "rsi_norm":         _safe(sig.rsi_norm, 3),
        "mfi_norm":         _safe(sig.mfi_norm, 3),
        # ── Method I Short-Side (Ch.16) ──
        "short_signal":         sig.short_signal,
        "cond_short_squeeze":   sig.cond_short_squeeze,
        "cond_short_price":     sig.cond_short_price,
        "cond_short_volume":    sig.cond_short_volume,
        "cond_short_ii_neg":    sig.cond_short_ii_neg,
        "cond_short_mfi_low":   sig.cond_short_mfi_low,
    }


def _fund_dict(fd):
    """Convert FundamentalData to a clean JSON-serialisable dict."""
    if fd is None or fd.fetch_error:
        return {"error": fd.fetch_error if fd else "No data"}

    def _qr(q):
        return {
            "period": q.period,
            "revenue": _safe(q.revenue), "revenue_str": _fmt_cr(q.revenue),
            "net_income": _safe(q.net_income), "net_income_str": _fmt_cr(q.net_income),
            "eps": _safe(q.eps), "ebitda": _safe(q.ebitda), "ebitda_str": _fmt_cr(q.ebitda),
            "gross_profit": _safe(q.gross_profit), "gross_profit_str": _fmt_cr(q.gross_profit),
            "operating_income": _safe(q.operating_income), "operating_income_str": _fmt_cr(q.operating_income),
            "interest_expense": _safe(q.interest_expense),
            "pretax_income": _safe(q.pretax_income),
            "tax_provision": _safe(q.tax_provision),
            "gross_margin": _safe(q.gross_margin),
            "operating_margin": _safe(q.operating_margin),
            "net_margin": _safe(q.net_margin),
        }

    def _sh(s):
        return {
            "period": s.period,
            "promoter": _safe(s.promoter, 1),
            "fii": _safe(s.fii, 1),
            "dii": _safe(s.dii, 1),
            "public": _safe(s.public, 1),
        }

    def _div(d):
        return {"date": d.date, "amount": d.amount}

    def _ed(e):
        return {
            "date": e.date,
            "eps_estimate": _safe(e.eps_estimate, 2) if e.eps_estimate is not None else None,
            "eps_reported": _safe(e.eps_reported, 2) if e.eps_reported is not None else None,
            "surprise_pct": _safe(e.surprise_pct, 2) if e.surprise_pct is not None else None,
        }

    return {
        "company_name":     fd.company_name,
        "sector":           fd.sector,
        "industry":         fd.industry,
        "exchange":         fd.exchange,
        "market_cap_str":   fd.market_cap_str,
        "ev_str":           fd.enterprise_value_str,
        "beta":             _safe(fd.beta),
        "current_price":    _safe(fd.current_price),
        "week_52_high":     _safe(fd.week_52_high),
        "week_52_low":      _safe(fd.week_52_low),
        "week_52_pct":      _safe(fd.week_52_pct),
        "description":      fd.description or "",
        "outstanding_shares": fd.outstanding_shares,
        # Valuation
        "pe_ratio":         _safe(fd.pe_ratio),
        "forward_pe":       _safe(fd.forward_pe),
        "pb_ratio":         _safe(fd.pb_ratio),
        "ps_ratio":         _safe(fd.ps_ratio),
        "ev_ebitda":        _safe(fd.ev_ebitda),
        "peg_ratio":        _safe(fd.peg_ratio),
        "price_to_fcf":     _safe(fd.price_to_fcf),
        "earning_yield":    _safe(fd.earning_yield),
        "book_value":       _safe(fd.book_value),
        "graham_number":    _safe(fd.graham_number),
        "price_to_intrinsic": _safe(fd.price_to_intrinsic, 3),
        # Profitability
        "roe":              _safe(fd.roe),
        "roa":              _safe(fd.roa),
        "roce":             _safe(fd.roce),
        "profit_margin":    _safe(fd.profit_margin),
        "operating_margin": _safe(fd.operating_margin),
        "gross_margin":     _safe(fd.gross_margin),
        "ebitda_margin":    _safe(fd.ebitda_margin),
        "ebitda":           _safe(fd.ebitda),
        "ebitda_str":       _fmt_cr(fd.ebitda),
        "eps_ttm":          _safe(fd.eps_ttm),
        "eps_forward":      _safe(fd.eps_forward),
        # Growth
        "revenue_growth":   _safe(fd.revenue_growth),
        "earnings_growth":  _safe(fd.earnings_growth),
        "total_revenue":    _safe(fd.total_revenue),
        "total_revenue_str": _fmt_cr(fd.total_revenue),
        "gross_profit_str": _fmt_cr(fd.gross_profit),
        "net_income":       _safe(fd.net_income),
        "net_income_str":   _fmt_cr(fd.net_income),
        "free_cash_flow":   _safe(fd.free_cash_flow),
        "fcf_str":          _fmt_cr(fd.free_cash_flow),
        "total_assets":     _safe(fd.total_assets),
        "total_assets_str": _fmt_cr(fd.total_assets),
        # Stability
        "debt_to_equity":   _safe(fd.debt_to_equity),
        "debt_to_ebitda":   _safe(fd.debt_to_ebitda),
        "current_ratio":    _safe(fd.current_ratio),
        "quick_ratio":      _safe(fd.quick_ratio),
        "cash_ratio":       _safe(fd.cash_ratio),
        "total_debt":       _safe(fd.total_debt),
        "total_debt_str":   _fmt_cr(fd.total_debt),
        "total_cash":       _safe(fd.total_cash),
        "total_cash_str":   _fmt_cr(fd.total_cash),
        "shareholders_equity": _safe(fd.shareholders_equity),
        "equity_str":       _fmt_cr(fd.shareholders_equity),
        "altman_z_score":   _safe(fd.altman_z_score),
        "interest_coverage": _safe(fd.interest_coverage),
        "asset_turnover":    _safe(fd.asset_turnover),
        "debt_to_assets":    _safe(fd.debt_to_assets, 4),
        # Dividends
        "dividend_yield":   _safe(fd.dividend_yield),
        "dividend_rate":    _safe(fd.dividend_rate),
        "payout_ratio":     _safe(fd.payout_ratio),
        "ex_dividend_date": fd.ex_dividend_date,
        "dividend_history": [_div(d) for d in (fd.dividend_history or [])],
        # Shareholding
        "promoter_holding": _safe(fd.promoter_holding, 1),
        "fii_holding":      _safe(fd.fii_holding, 1),
        "dii_holding":      _safe(fd.dii_holding, 1),
        "public_holding":   _safe(fd.public_holding, 1),
        "float_pct":        _safe(fd.float_pct, 1),
        # History
        "shareholding_history": [_sh(s) for s in (fd.shareholding_history or [])],
        "quarterly_results":    [_qr(q) for q in (fd.quarterly_results or [])],
        # Earnings Calendar
        "upcoming_results_date": fd.upcoming_results_date,
        "earnings_estimate_eps": _safe(fd.earnings_estimate_eps, 2) if fd.earnings_estimate_eps else None,
        "earnings_dates_history": [_ed(e) for e in (fd.earnings_dates_history or [])],
        # Scores
        "valuation_score":      fd.valuation_score,
        "profitability_score":  fd.profitability_score,
        "growth_score":         fd.growth_score,
        "stability_score":      fd.stability_score,
        "fundamental_score":    fd.fundamental_score,
        "fundamental_verdict":  fd.fundamental_verdict,
        # Verdict signal
        "fundamental_signal":   fd.fundamental_signal or "",
        "signal_strength":      fd.signal_strength or "",
        "signal_color":         fd.signal_color or "",
        # Text
        "valuation_analysis":     fd.valuation_analysis,
        "profitability_analysis": fd.profitability_analysis,
        "growth_analysis":        fd.growth_analysis,
        "stability_analysis":     fd.stability_analysis,
        "conviction_message":     fd.conviction_message,
        # Descriptive Summaries
        "quarterly_analysis":     fd.quarterly_analysis,
        "shareholding_verdict":   fd.shareholding_verdict,
        "dividend_summary":       fd.dividend_summary,
        # Delivery Data
        "delivery_quantity":      fd.delivery_quantity,
        "traded_quantity":        fd.traded_quantity,
        "delivery_pct":           _safe(fd.delivery_pct) if fd.delivery_pct is not None else None,
        "delivery_date":          fd.delivery_date,
        "delivery_analysis":      fd.delivery_analysis,
        # Historical Delivery
        "delivery_history": [
            {
                "date": dh.date, "close_price": _safe(dh.close_price),
                "volume": dh.volume, "delivery_qty": dh.delivery_qty,
                "delivery_pct": _safe(dh.delivery_pct),
            }
            for dh in (fd.delivery_history or [])
        ],
        "delivery_history_analysis": fd.delivery_history_analysis,
        # Bulk/Block Deals & Insider Trading
        "bulk_block_deals":       [
            {
                "date": d.date, "deal_type": d.deal_type,
                "client_name": d.client_name, "buy_sell": d.buy_sell,
                "quantity": d.quantity, "price": _safe(d.price),
                "remarks": d.remarks,
            }
            for d in (fd.bulk_block_deals or [])
        ],
        "insider_trades":         [
            {
                "person_name": t.person_name, "category": t.category,
                "txn_type": t.txn_type, "shares": t.shares,
                "value": t.value, "date_from": t.date_from,
                "date_to": t.date_to, "post_shares": t.post_shares,
                "mode": t.mode,
            }
            for t in (fd.insider_trades or [])
        ],
        "deals_analysis":         fd.deals_analysis,
        # Financial Statements
        "annual_financials": [
            {
                "year": af.year,
                "total_revenue": _safe(af.total_revenue),
                "cost_of_revenue": _safe(af.cost_of_revenue),
                "gross_profit": _safe(af.gross_profit),
                "operating_expense": _safe(af.operating_expense),
                "operating_income": _safe(af.operating_income),
                "ebitda": _safe(af.ebitda),
                "ebit": _safe(af.ebit),
                "interest_expense": _safe(af.interest_expense),
                "pretax_income": _safe(af.pretax_income),
                "tax_provision": _safe(af.tax_provision),
                "net_income": _safe(af.net_income),
                "diluted_eps": _safe(af.diluted_eps),
                "total_assets": _safe(af.total_assets),
                "total_current_assets": _safe(af.total_current_assets),
                "cash_and_equivalents": _safe(af.cash_and_equivalents),
                "inventory": _safe(af.inventory),
                "accounts_receivable": _safe(af.accounts_receivable),
                "net_ppe": _safe(af.net_ppe),
                "goodwill": _safe(af.goodwill),
                "total_liabilities": _safe(af.total_liabilities),
                "total_current_liab": _safe(af.total_current_liab),
                "long_term_debt": _safe(af.long_term_debt),
                "total_debt": _safe(af.total_debt),
                "stockholders_equity": _safe(af.stockholders_equity),
                "retained_earnings": _safe(af.retained_earnings),
                "working_capital": _safe(af.working_capital),
                "operating_cashflow": _safe(af.operating_cashflow),
                "capex": _safe(af.capex),
                "free_cashflow": _safe(af.free_cashflow),
                "investing_cashflow": _safe(af.investing_cashflow),
                "financing_cashflow": _safe(af.financing_cashflow),
                "dividends_paid": _safe(af.dividends_paid),
                "debt_repayment": _safe(af.debt_repayment),
                "debt_issuance": _safe(af.debt_issuance),
                "depreciation": _safe(af.depreciation),
                "gross_margin": _safe(af.gross_margin),
                "operating_margin": _safe(af.operating_margin),
                "net_margin": _safe(af.net_margin),
                "roe": _safe(af.roe),
                "roa": _safe(af.roa),
                "debt_to_equity": _safe(af.debt_to_equity, 3),
                "current_ratio": _safe(af.current_ratio),
                "interest_coverage": _safe(af.interest_coverage),
                "fcf_margin": _safe(af.fcf_margin),
            }
            for af in (fd.annual_financials or [])
        ],
        "financial_statements_analysis": fd.financial_statements_analysis,
    }


# ─────────────────────────────────────────────────────────────────
#  ROUTES
# ─────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Main landing page."""
    return render_template("index.html")


@app.route("/analyze/<ticker_raw>")
def analyze_page(ticker_raw):
    """Full analysis page for a single stock."""
    ticker = normalise_ticker(ticker_raw)
    return render_template("analyze.html", ticker=ticker)


@app.route("/api/tickers")
def api_tickers():
    """Return list of all available tickers."""
    tickers = get_all_tickers_from_csv(CSV_DIR)
    return jsonify(tickers)


@app.route("/api/analyze/<ticker_raw>")
def api_analyze(ticker_raw):
    """Full analysis JSON — signals + fundamentals + chart data."""
    ticker = normalise_ticker(ticker_raw)

    # Load price data
    df = load_stock_data(ticker, csv_dir=CSV_DIR)
    if df is None:
        from bb_squeeze.data_loader import fetch_live_data
        df = fetch_live_data(ticker)
    if df is None:
        return jsonify({"error": f"No data for {ticker}"}), 404

    # Compute indicators
    df = compute_all_indicators(df)

    # Generate signals
    sig = analyze_signals(ticker, df)

    # Run additional strategies (Methods II, III, IV)
    strategies = run_all_strategies(df)

    # Fetch fundamentals
    fd = None
    try:
        fd = fetch_fundamentals(ticker)
    except Exception:
        pass

    # Build chart data
    chart = _build_chart_data(df)

    # Quantitative Trading Strategy (separate system)
    quant = None
    try:
        quant = run_quant_analysis(df)
    except Exception:
        pass

    # Triple Conviction Analysis (BB + TA + PA cross-validated)
    triple = None
    try:
        triple = run_triple_analysis(df, ticker=ticker)
    except Exception:
        pass

    # Price Action (Al Brooks)
    pa = None
    try:
        bb_data_for_pa = {
            "buy_signal": sig.buy_signal,
            "sell_signal": sig.sell_signal,
            "direction_lean": sig.direction_lean,
            "confidence": sig.confidence,
            "phase": sig.phase,
        }
        ta_data_for_pa = triple.get("ta_signal") if triple else None
        pa_raw = run_price_action_analysis(
            df=df, ticker=ticker,
            bb_data=bb_data_for_pa, ta_data=ta_data_for_pa,
            hybrid_data=triple,
        )
        pa = pa_result_to_dict(pa_raw)
    except Exception:
        pass

    # Data freshness
    freshness = get_data_freshness(df)

    return jsonify({
        "signal":       _signal_dict(sig),
        "strategies":   [strategy_result_to_dict(sr) for sr in strategies],
        "fundamentals": _fund_dict(fd),
        "chart":        chart,
        "quant":        quant,
        "hybrid":       triple,
        "triple":       triple,
        "pa":           pa,
        "data_freshness": freshness,
    })


@app.route("/api/refresh-data/<ticker_raw>", methods=["POST"])
def api_refresh_data(ticker_raw):
    """Force fresh data download from Yahoo Finance for a ticker."""
    from bb_squeeze.data_loader import fetch_live_data
    ticker = normalise_ticker(ticker_raw)
    df = fetch_live_data(ticker, period="2y")
    if df is None:
        return jsonify({"error": f"Could not fetch live data for {ticker}"}), 404

    # Save fresh data to CSV
    csv_path = os.path.join(CSV_DIR, f"{ticker}.csv")
    df.to_csv(csv_path)

    freshness = get_data_freshness(df)
    return jsonify({
        "success": True,
        "ticker": ticker,
        "rows": len(df),
        "data_freshness": freshness,
    })


@app.route("/api/scan")
def api_scan():
    """Quick scan — returns signals for all stocks (no fundamentals)."""
    tickers = get_all_tickers_from_csv(CSV_DIR)
    results = []
    for t in tickers:
        try:
            df = load_stock_data(t, csv_dir=CSV_DIR, use_live_fallback=False)
            if df is None or len(df) < 50:
                continue
            df = compute_all_indicators(df)
            sig = analyze_signals(t, df)
            if sig.phase in ("INSUFFICIENT_DATA", "ERROR"):
                continue
            results.append(_signal_dict(sig))
        except Exception:
            continue

    return jsonify(results)


@app.route("/api/scan/strategies")
def api_scan_strategies():
    """Scan all stocks with ALL 4 methods — returns combined results."""
    tickers = get_all_tickers_from_csv(CSV_DIR)
    results = []
    for t in tickers:
        try:
            df = load_stock_data(t, csv_dir=CSV_DIR, use_live_fallback=False)
            if df is None or len(df) < 50:
                continue
            df = compute_all_indicators(df)

            # Method I (Squeeze)
            sig = analyze_signals(t, df)
            if sig.phase in ("INSUFFICIENT_DATA", "ERROR"):
                continue
            m1 = _signal_dict(sig)

            # Methods II, III, IV
            strats = run_all_strategies(df)
            strat_dicts = [strategy_result_to_dict(sr) for sr in strats]

            results.append({
                "ticker":    t,
                "price":     m1["current_price"],
                "m1":        m1,
                "strategies": strat_dicts,
            })
        except Exception:
            continue

    return jsonify(results)


@app.route("/api/chart/<ticker_raw>")
def api_chart(ticker_raw):
    """Chart data at different intervals: 15m, daily (default), weekly."""
    ticker = normalise_ticker(ticker_raw)
    interval = request.args.get("interval", "daily")

    if interval == "15m":
        import yfinance as yf
        try:
            yf_ticker = yf.Ticker(ticker)
            df = yf_ticker.history(period="5d", interval="15m")
            if df is None or len(df) < 5:
                return jsonify({"error": "No intraday data available"}), 404
            # Unix timestamps (seconds) for Lightweight Charts intraday
            timestamps = [int(d.timestamp()) for d in df.index]
            return jsonify({
                "interval": "15m",
                "dates": timestamps,
                "open":   [round(float(v), 2) for v in df["Open"]],
                "high":   [round(float(v), 2) for v in df["High"]],
                "low":    [round(float(v), 2) for v in df["Low"]],
                "close":  [round(float(v), 2) for v in df["Close"]],
                "volume": [int(v) for v in df["Volume"]],
            })
        except Exception:
            return jsonify({"error": "Failed to fetch intraday data"}), 500

    elif interval == "weekly":
        df = load_stock_data(ticker, csv_dir=CSV_DIR)
        if df is None:
            from bb_squeeze.data_loader import fetch_live_data
            df = fetch_live_data(ticker)
        if df is None:
            return jsonify({"error": f"No data for {ticker}"}), 404

        # Resample daily → weekly OHLCV
        weekly = df.resample("W").agg({
            "Open": "first", "High": "max", "Low": "min",
            "Close": "last", "Volume": "sum",
        }).dropna(subset=["Close"])

        # Compute 20-week Bollinger Bands
        weekly["BB_Mid"] = weekly["Close"].rolling(20).mean()
        bb_std = weekly["Close"].rolling(20).std()
        weekly["BB_Upper"] = weekly["BB_Mid"] + 2 * bb_std
        weekly["BB_Lower"] = weekly["BB_Mid"] - 2 * bb_std

        weekly = weekly.tail(104)  # ~2 years of weekly candles
        dates = [d.strftime("%Y-%m-%d") for d in weekly.index]

        def _ws(v):
            return round(float(v), 2) if pd.notna(v) else None

        return jsonify({
            "interval": "weekly",
            "dates":    dates,
            "open":     [_ws(v) for v in weekly["Open"]],
            "high":     [_ws(v) for v in weekly["High"]],
            "low":      [_ws(v) for v in weekly["Low"]],
            "close":    [_ws(v) for v in weekly["Close"]],
            "volume":   [int(v) for v in weekly["Volume"]],
            "bb_upper": [_ws(v) for v in weekly["BB_Upper"]],
            "bb_mid":   [_ws(v) for v in weekly["BB_Mid"]],
            "bb_lower": [_ws(v) for v in weekly["BB_Lower"]],
        })

    else:
        return jsonify({"error": "Invalid interval. Use: 15m, daily, weekly"}), 400


# ─────────────────────────────────────────────────────────────────
#  HISTORICAL DATA DOWNLOAD — Multi-threaded
# ─────────────────────────────────────────────────────────────────

# Shared download state
_dl_state = {
    "running": False,
    "total": 0,
    "done": 0,
    "saved": 0,
    "skipped": 0,
    "failed": 0,
    "current": "",
    "finished": False,
    "error": None,
    "start_time": 0,
}
_dl_lock = threading.Lock()


def _download_worker(ticker, start, end, save_path, staleness_days):
    """Download one ticker. Returns (ticker, status_str)."""
    import yfinance as yf
    file_path = os.path.join(save_path, f"{ticker}.csv")

    # Check if file exists and is fresh
    if os.path.exists(file_path):
        try:
            existing = pd.read_csv(file_path)
            if existing.columns[0] == "Date" and len(existing) > 10:
                last_date = pd.to_datetime(existing["Date"].iloc[-1])
                today = pd.Timestamp.today().normalize()
                if (today - last_date).days <= staleness_days:
                    return (ticker, "skipped")
        except Exception:
            pass  # corrupt → re-download

    for attempt in range(2):
        try:
            t = yf.Ticker(ticker)
            raw = t.history(start=start, end=end, auto_adjust=False)
            if raw is None or raw.empty:
                if attempt == 0:
                    _time.sleep(1)
                    continue
                return (ticker, "no_data")

            df = raw.copy()
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [col[0] if isinstance(col, tuple) else col
                              for col in df.columns]
            wanted = ["Open", "High", "Low", "Close", "Adj Close", "Volume"]
            df = df[[c for c in wanted if c in df.columns]]
            if "Close" not in df.columns:
                return (ticker, "no_data")

            dt_index = pd.to_datetime(df.index, utc=True).tz_convert("Asia/Kolkata")
            df.index = dt_index.normalize().tz_localize(None).date
            df.index = pd.DatetimeIndex(df.index)
            df.index.name = "Date"
            df = df.sort_index()
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            df = df.dropna(subset=["Close"])
            df = df[df["Close"] > 0]

            if df.empty:
                return (ticker, "no_data")

            df.to_csv(file_path)
            return (ticker, "saved")
        except Exception:
            if attempt == 0:
                _time.sleep(2)
            else:
                return (ticker, "error")
    return (ticker, "error")


def _run_download_thread(tickers, save_path, start_date, end_date,
                         staleness_days, max_workers):
    """Background thread that runs the multi-threaded download."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    global _dl_state

    os.makedirs(save_path, exist_ok=True)
    valid = [t for t in tickers if t and not t.startswith(".") and len(t) > 3]

    with _dl_lock:
        _dl_state["total"] = len(valid)
        _dl_state["done"] = 0
        _dl_state["saved"] = 0
        _dl_state["skipped"] = 0
        _dl_state["failed"] = 0
        _dl_state["current"] = ""
        _dl_state["finished"] = False
        _dl_state["error"] = None
        _dl_state["start_time"] = _time.time()

    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(
                    _download_worker, t, start_date, end_date,
                    save_path, staleness_days
                ): t for t in valid
            }
            for future in as_completed(futures):
                ticker = futures[future]
                try:
                    _, status = future.result()
                except Exception:
                    status = "error"

                with _dl_lock:
                    _dl_state["done"] += 1
                    _dl_state["current"] = ticker
                    if status == "saved":
                        _dl_state["saved"] += 1
                    elif status == "skipped":
                        _dl_state["skipped"] += 1
                    else:
                        _dl_state["failed"] += 1
    except Exception as e:
        with _dl_lock:
            _dl_state["error"] = str(e)
    finally:
        with _dl_lock:
            _dl_state["running"] = False
            _dl_state["finished"] = True


# ── Ticker list management ────────────────────────────────────────

@app.route("/api/tickers/refresh", methods=["POST"])
def api_tickers_refresh():
    """Fetch latest ticker list from NSE and update cache."""
    try:
        from historical_data import refresh_tickers_cache, TICKERS_CACHE_FILE
        body = request.get_json(silent=True) or {}
        include_sme = body.get("include_sme", False)
        result = refresh_tickers_cache(include_sme=include_sme)
        return jsonify({
            "status": "ok",
            "count": result["count"],
            "fetched_at": result["fetched_at"],
            "source": result["source"],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tickers/info")
def api_tickers_info():
    """Return info about the current ticker cache."""
    from historical_data import TICKERS_CACHE_FILE, TICKERS
    import json as _json
    info = {"count": len(TICKERS), "source": "hardcoded fallback", "fetched_at": None}
    if os.path.exists(TICKERS_CACHE_FILE):
        try:
            with open(TICKERS_CACHE_FILE) as f:
                data = _json.load(f)
            info["count"] = data.get("count", len(TICKERS))
            info["source"] = data.get("source", "cache")
            info["fetched_at"] = data.get("fetched_at")
            info["include_sme"] = data.get("include_sme", False)
        except Exception:
            pass
    return jsonify(info)


# ── EOD data availability check ──────────────────────────────────

_eod_cache = {"ts": 0, "data": None}  # cached YF probe (avoids rate limits)
_EOD_CACHE_TTL = 900  # 15 minutes

@app.route("/api/eod-status")
def api_eod_status():
    """Check whether new EOD data is available for download.

    Compares the latest local CSV date against what Yahoo Finance
    actually has (probed & cached for 15 min).  Only says "Available"
    when YFinance truly has newer data than local CSVs.
    """
    from bb_squeeze.data_loader import _last_expected_trading_date, _is_nse_trading_day
    from datetime import datetime, timedelta
    import pytz

    ist = pytz.timezone("Asia/Kolkata")
    now_ist = datetime.now(ist)

    # Market closes at 15:30 IST; data appears on Yahoo ~16:00-16:30
    market_closed = now_ist.hour > 16 or (now_ist.hour == 16 and now_ist.minute >= 15)
    is_trading_day = _is_nse_trading_day(now_ist.replace(tzinfo=None))

    # ── Find latest local CSV date (sample 5 liquid stocks) ──
    sample_tickers = ["RELIANCE.NS", "TCS.NS", "INFY.NS", "HDFCBANK.NS", "SBIN.NS"]
    local_dates = []
    for ticker in sample_tickers:
        csv_path = os.path.join(CSV_DIR, f"{ticker}.csv")
        if not os.path.exists(csv_path):
            continue
        try:
            df = pd.read_csv(csv_path, usecols=["Date"], parse_dates=["Date"])
            if not df.empty:
                local_dates.append(df["Date"].iloc[-1].date())
        except Exception:
            pass

    local_last = max(local_dates) if local_dates else None

    # ── Probe YFinance for latest available date (cached) ──
    yf_last = None
    now_ts = _time.time()
    if _eod_cache["data"] and (now_ts - _eod_cache["ts"]) < _EOD_CACHE_TTL:
        yf_last = _eod_cache["data"]
    else:
        try:
            import yfinance as yf
            t = yf.Ticker("RELIANCE.NS")
            hist = t.history(period="5d", auto_adjust=False)
            if hist is not None and not hist.empty:
                yf_last = hist.index[-1].date()
                _eod_cache["data"] = yf_last
                _eod_cache["ts"] = now_ts
        except Exception:
            # YF probe failed — fall back to expected-date heuristic
            last_td = _last_expected_trading_date()
            yf_last = last_td.date() if hasattr(last_td, "date") else last_td

    # ── Decide availability ──
    if local_last is None:
        available = True
        reason = "No local data — download recommended"
    elif yf_last and local_last < yf_last:
        available = True
        reason = f"Local data: {local_last} → YFinance has: {yf_last}"
    elif is_trading_day and not market_closed:
        available = False
        reason = f"Market still open — local data through {local_last}"
    else:
        available = False
        reason = f"Up to date (through {local_last})"

    return jsonify({
        "available": available,
        "reason": reason,
        "local_last_date": str(local_last) if local_last else None,
        "yf_last_date": str(yf_last) if yf_last else None,
        "market_open_today": is_trading_day and not market_closed,
    })


@app.route("/api/download/start", methods=["POST"])
def api_download_start():
    """Start historical data download in background."""
    global _dl_state

    with _dl_lock:
        if _dl_state["running"]:
            return jsonify({"error": "Download already in progress"}), 409

    # Import config from historical_data.py; reload tickers fresh
    from historical_data import load_tickers, START_DATE, SAVE_PATH, STALENESS_DAYS
    tickers = load_tickers()

    body = request.get_json(silent=True) or {}
    force = body.get("force", False)
    max_workers = min(int(body.get("threads", 8)), 20)  # cap at 20

    # yfinance 'end' is exclusive, so use tomorrow to include today's data
    end_date = (date.today() + timedelta(days=1)).strftime("%Y-%m-%d")
    staleness = 0 if force else STALENESS_DAYS

    with _dl_lock:
        _dl_state["running"] = True
        _dl_state["finished"] = False

    thread = threading.Thread(
        target=_run_download_thread,
        args=(tickers, SAVE_PATH, START_DATE, end_date, staleness, max_workers),
        daemon=True,
    )
    thread.start()

    return jsonify({
        "status": "started",
        "total": len([t for t in tickers if t and not t.startswith(".") and len(t) > 3]),
        "threads": max_workers,
        "force": force,
    })


@app.route("/api/download/status")
def api_download_status():
    """SSE stream of download progress."""
    def generate():
        while True:
            with _dl_lock:
                state = dict(_dl_state)
            elapsed = _time.time() - state["start_time"] if state["start_time"] else 0
            data = json.dumps({
                "running": state["running"],
                "total": state["total"],
                "done": state["done"],
                "saved": state["saved"],
                "skipped": state["skipped"],
                "failed": state["failed"],
                "current": state["current"],
                "finished": state["finished"],
                "error": state["error"],
                "elapsed": round(elapsed, 1),
            })
            yield f"data: {data}\n\n"

            if state["finished"]:
                break
            _time.sleep(0.5)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/download/state")
def api_download_state():
    """Quick poll of current download state (non-SSE)."""
    with _dl_lock:
        state = dict(_dl_state)
    elapsed = _time.time() - state["start_time"] if state["start_time"] else 0
    state["elapsed"] = round(elapsed, 1)
    return jsonify(state)


# ─────────────────────────────────────────────────────────────────
#  TRADE P&L DASHBOARD ROUTES
# ─────────────────────────────────────────────────────────────────

@app.route("/trades")
def trades_page():
    return render_template("trades.html")


@app.route("/api/trades", methods=["GET"])
def api_get_trades():
    """Return all trades with calculated P&L."""
    rows = get_all_trades(user_id=_uid(), is_admin=_is_admin())
    results = []
    for t in rows:
        pnl = calculate_trade(
            stock=t["stock"], platform=t["platform"],
            trade_type=t["trade_type"], exchange=t["exchange"],
            quantity=t["quantity"], buy_price=t["buy_price"],
            sell_price=t["sell_price"], buy_date=t["buy_date"],
            sell_date=t["sell_date"],
        )
        results.append({**t, "pnl": pnl.to_dict()})
    return jsonify(results)


@app.route("/api/trades", methods=["POST"])
def api_add_trade():
    """Add a new trade."""
    data = request.get_json(force=True)
    required = ["stock", "quantity", "buy_price", "sell_price", "buy_date", "sell_date"]
    missing = [k for k in required if not data.get(k)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    tid = add_trade(data, user_id=_uid())
    return jsonify({"id": tid, "status": "ok"})


@app.route("/api/trades/<int:tid>", methods=["PUT"])
def api_update_trade(tid):
    """Update an existing trade."""
    data = request.get_json(force=True)
    ok = update_trade(tid, data, user_id=_uid(), is_admin=_is_admin())
    return jsonify({"status": "ok" if ok else "not_found"})


@app.route("/api/trades/<int:tid>", methods=["DELETE"])
def api_delete_trade(tid):
    """Delete a trade. If it was auto-logged from a portfolio position, also delete that position."""
    uid = _uid()
    is_admin = _is_admin()
    existing = get_trade(tid, user_id=uid, is_admin=is_admin)
    linked_pid = (existing or {}).get("position_id")
    ok = delete_trade(tid, user_id=uid, is_admin=is_admin)
    position_deleted = False
    if ok and linked_pid:
        try:
            position_deleted = delete_position(int(linked_pid), user_id=uid, is_admin=is_admin)
        except Exception:
            position_deleted = False
    return jsonify({"status": "ok" if ok else "not_found", "position_deleted": position_deleted})


@app.route("/api/trades/summary")
def api_trades_summary():
    """FY-wise tax summary with LTCG exemption applied."""
    rows = get_all_trades(user_id=_uid(), is_admin=_is_admin())
    items = []
    for t in rows:
        pnl = calculate_trade(
            stock=t["stock"], platform=t["platform"],
            trade_type=t["trade_type"], exchange=t["exchange"],
            quantity=t["quantity"], buy_price=t["buy_price"],
            sell_price=t["sell_price"], buy_date=t["buy_date"],
            sell_date=t["sell_date"],
        )
        items.append({"sell_date": t["sell_date"], "pnl": pnl.to_dict()})
    summary = calculate_fy_summary(items)
    return jsonify(summary)


def _build_trades_with_pnl():
    """Shared helper: pull trades for the current user with full P&L attached."""
    rows = get_all_trades(user_id=_uid(), is_admin=_is_admin())
    enriched = []
    for t in rows:
        pnl = calculate_trade(
            stock=t["stock"], platform=t["platform"],
            trade_type=t["trade_type"], exchange=t["exchange"],
            quantity=t["quantity"], buy_price=t["buy_price"],
            sell_price=t["sell_price"], buy_date=t["buy_date"],
            sell_date=t["sell_date"],
        )
        enriched.append({**t, "pnl": pnl.to_dict()})
    return enriched


def _current_user_display_name():
    """Best-effort display name for PDF cover line."""
    u = getattr(g, "user", None)
    if not u:
        return None
    full = " ".join(filter(None, [u.get("first_name"), u.get("last_name")])).strip()
    return full or u.get("username") or u.get("email")


@app.route("/api/trades/export/pdf")
def api_trades_export_pdf():
    """Download a comprehensive Trade History PDF report."""
    from web.pdf_trades import build_trade_history_pdf

    trades = _build_trades_with_pnl()
    pdf_bytes = build_trade_history_pdf(trades, user_name=_current_user_display_name())
    filename = f"Hiranya_Trade_History_{date.today().isoformat()}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


@app.route("/api/trades/summary/export/pdf")
def api_trades_summary_export_pdf():
    """Download an ITR-ready Financial Year Tax Summary PDF."""
    from web.pdf_trades import build_fy_tax_summary_pdf

    trades = _build_trades_with_pnl()
    items = [{"sell_date": t["sell_date"], "pnl": t["pnl"]} for t in trades]
    fy_summary = calculate_fy_summary(items)
    pdf_bytes = build_fy_tax_summary_pdf(
        fy_summary, trades=trades, user_name=_current_user_display_name(),
    )
    filename = f"Hiranya_FY_Tax_Summary_{date.today().isoformat()}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


# ─────────────────────────────────────────────────────────────────
#  PORTFOLIO TRACKER (completely isolated)
# ─────────────────────────────────────────────────────────────────

@app.route("/portfolio")
def portfolio_page():
    return render_template("portfolio.html")


@app.route("/api/portfolio", methods=["GET"])
def api_portfolio_list():
    """List all portfolio positions."""
    uid, adm = _uid(), _is_admin()
    filt = request.args.get("filter", "all")
    if filt == "open":
        positions = get_open_positions(user_id=uid, is_admin=adm)
    elif filt == "closed":
        positions = get_closed_positions(user_id=uid, is_admin=adm)
    else:
        positions = get_all_positions(user_id=uid, is_admin=adm)
    return jsonify(positions)


@app.route("/api/portfolio", methods=["POST"])
def api_portfolio_add():
    """Add a new portfolio position."""
    data = request.get_json(force=True)
    required = ["ticker", "strategy_code", "buy_price", "buy_date"]
    for f in required:
        if f not in data or not data[f]:
            return jsonify({"error": f"Missing required field: {f}"}), 400
    if data["strategy_code"].upper() not in ("M1", "M2", "M3", "M4"):
        return jsonify({"error": "strategy_code must be M1, M2, M3, or M4"}), 400
    pid = add_position(data, user_id=_uid())
    return jsonify({"status": "ok", "id": pid})


@app.route("/api/portfolio/<int:pid>", methods=["GET"])
def api_portfolio_get(pid):
    """Get a single position."""
    pos = get_position(pid, user_id=_uid(), is_admin=_is_admin())
    if not pos:
        return jsonify({"error": "not_found"}), 404
    return jsonify(pos)


@app.route("/api/portfolio/<int:pid>", methods=["PUT"])
def api_portfolio_update(pid):
    """Update a position's details."""
    data = request.get_json(force=True)
    ok = update_position(pid, data, user_id=_uid(), is_admin=_is_admin())
    return jsonify({"status": "ok" if ok else "not_found"})


@app.route("/api/portfolio/<int:pid>", methods=["DELETE"])
def api_portfolio_delete(pid):
    """Delete a position. Also remove any auto-logged trade entries that point to this position."""
    uid = _uid()
    is_admin = _is_admin()
    ok = delete_position(pid, user_id=uid, is_admin=is_admin)
    trades_deleted = 0
    if ok:
        try:
            trades_deleted = delete_trades_by_position(pid, user_id=uid, is_admin=is_admin)
        except Exception:
            trades_deleted = 0
    return jsonify({"status": "ok" if ok else "not_found", "trades_deleted": trades_deleted})


@app.route("/api/portfolio/<int:pid>/close", methods=["POST"])
def api_portfolio_close(pid):
    """Close an open position with sell details and auto-log a trade entry."""
    data = request.get_json(force=True)
    sell_price = data.get("sell_price")
    sell_date = data.get("sell_date")
    if not sell_price or not sell_date:
        return jsonify({"error": "sell_price and sell_date required"}), 400

    uid = _uid()
    is_admin = _is_admin()

    pos = get_position(pid, user_id=uid, is_admin=is_admin)
    if not pos or pos.get("status") != "OPEN":
        return jsonify({"status": "not_found_or_already_closed"})

    ok = close_position(pid, float(sell_price), sell_date, data.get("sell_reason", ""),
                        user_id=uid, is_admin=is_admin)
    if not ok:
        return jsonify({"status": "not_found_or_already_closed"})

    trade_id = None
    try:
        ticker_clean = (pos.get("ticker") or "").upper().replace(".NS", "").strip()
        reason = (data.get("sell_reason") or "").strip()
        strat = (pos.get("strategy_code") or "").strip()
        note_bits = [b for b in [f"Auto-closed from portfolio (#{pid})", strat, reason] if b]
        trade_payload = {
            "stock":      ticker_clean,
            "platform":   (data.get("platform") or "zerodha").lower(),
            "trade_type": (data.get("trade_type") or "delivery").lower(),
            "exchange":   (data.get("exchange") or "NSE").upper(),
            "quantity":   int(pos.get("quantity") or 1),
            "buy_price":  float(pos.get("buy_price")),
            "sell_price": float(sell_price),
            "buy_date":   pos.get("buy_date"),
            "sell_date":  sell_date,
            "notes":      " — ".join(note_bits),
            "position_id": pid,
        }
        trade_id = add_trade(trade_payload, user_id=uid)
    except Exception as exc:
        # Position is already closed; surface trade-log failure but don't roll back.
        return jsonify({"status": "ok", "trade_logged": False, "trade_error": str(exc)})

    return jsonify({"status": "ok", "trade_logged": True, "trade_id": trade_id})


@app.route("/api/portfolio/<int:pid>/reopen", methods=["POST"])
def api_portfolio_reopen(pid):
    """Reopen a closed position. Also removes any auto-logged trade entries since the trade no longer exists."""
    uid = _uid()
    is_admin = _is_admin()
    ok = reopen_position(pid, user_id=uid, is_admin=is_admin)
    trades_deleted = 0
    if ok:
        try:
            trades_deleted = delete_trades_by_position(pid, user_id=uid, is_admin=is_admin)
        except Exception:
            trades_deleted = 0
    return jsonify({"status": "ok" if ok else "not_found_or_already_open", "trades_deleted": trades_deleted})


@app.route("/api/portfolio/<int:pid>/analyze", methods=["GET"])
def api_portfolio_analyze(pid):
    """Full daily analysis for a single open position."""
    pos = get_position(pid, user_id=_uid(), is_admin=_is_admin())
    if not pos:
        return jsonify({"error": "not_found"}), 404
    result = analyze_position(pos)
    return jsonify(result)


@app.route("/api/portfolio/<int:pid>/analyze/pdf", methods=["GET"])
def api_portfolio_analyze_pdf(pid):
    """Render the single-position analysis as a clean A4 PDF."""
    from web.pdf_position import build_position_analysis_pdf

    pos = get_position(pid, user_id=_uid(), is_admin=_is_admin())
    if not pos:
        return jsonify({"error": "not_found"}), 404

    analysis = analyze_position(pos)
    pdf_bytes = build_position_analysis_pdf(analysis)
    ticker = (pos.get("ticker") or "position").upper().replace(" ", "_")
    filename = f"Position_Analysis_{ticker}_{date.today().isoformat()}.pdf"

    from flask import Response
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


@app.route("/api/portfolio/analyze-all", methods=["GET"])
def api_portfolio_analyze_all():
    """Analyze all open positions."""
    positions = get_open_positions(user_id=_uid(), is_admin=_is_admin())
    results = analyze_all_open_positions(positions)
    return jsonify(results)


@app.route("/api/portfolio/unrealized-daily", methods=["GET"])
def api_portfolio_unrealized_daily():
    """Return per-day unrealized P&L as total and per-stock series for open positions."""
    positions = get_open_positions(user_id=_uid(), is_admin=_is_admin())
    if not positions:
        return jsonify({"dates": [], "pnl": [], "series": [], "meta": {"positions": 0, "stocks": 0}})

    pnl_series = []

    for p in positions:
        ticker = (p.get("ticker") or "").strip().upper()
        if not ticker:
            continue

        df = load_stock_data(ticker, CSV_DIR, use_live_fallback=False)
        if df is None or df.empty or "Close" not in df.columns:
            continue

        try:
            buy_dt = pd.to_datetime(p.get("buy_date"))
            buy_price = float(p.get("buy_price") or 0)
            qty = int(p.get("quantity") or 1)
        except Exception:
            continue

        if buy_price <= 0 or qty <= 0:
            continue

        close = df["Close"].copy()
        close.index = pd.to_datetime(close.index)
        close = close.sort_index()
        close = close[close.index >= buy_dt]
        if close.empty:
            continue

        # Position-level unrealized P&L per day: (close - buy_price) * quantity.
        pnl = (close - buy_price) * qty
        pnl.name = ticker
        pnl_series.append((ticker, pnl))

    if not pnl_series:
        return jsonify({"dates": [], "pnl": [], "series": [], "meta": {"positions": len(positions), "stocks": 0}})

    all_dates = sorted(set().union(*[set(s.index) for _, s in pnl_series]))
    idx = pd.DatetimeIndex(all_dates)

    by_ticker = {}
    for ticker, s in pnl_series:
        aligned = s.reindex(idx).ffill().fillna(0.0)
        if ticker in by_ticker:
            by_ticker[ticker] = by_ticker[ticker] + aligned
        else:
            by_ticker[ticker] = aligned

    series_payload = []
    for ticker in sorted(by_ticker.keys()):
        ts = by_ticker[ticker]
        series_payload.append({
            "ticker": ticker,
            "label": ticker.replace(".NS", ""),
            "pnl": [round(float(v), 2) for v in ts.values],
            "latest": round(float(ts.iloc[-1]), 2) if len(ts) else 0.0,
        })

    portfolio_pnl = pd.concat(list(by_ticker.values()), axis=1).sum(axis=1)

    return jsonify({
        "dates": [d.strftime("%Y-%m-%d") for d in portfolio_pnl.index],
        "pnl": [round(float(v), 2) for v in portfolio_pnl.values],
        "series": series_payload,
        "meta": {
            "positions": len(positions),
            "stocks": len(series_payload),
            "series_used": len(pnl_series),
            "latest": round(float(portfolio_pnl.iloc[-1]), 2) if len(portfolio_pnl) else 0.0,
        },
    })


# ─────────────────────────────────────────────────────────────────
#  PORTFOLIO EXPORT (Excel / PDF)
# ─────────────────────────────────────────────────────────────────

def _days_between(date_str, end=None):
    """Helper: days between a YYYY-MM-DD string and end (or today)."""
    try:
        d1 = pd.to_datetime(date_str).date()
    except Exception:
        return 0
    if end:
        try:
            d2 = pd.to_datetime(end).date()
        except Exception:
            d2 = date.today()
    else:
        d2 = date.today()
    return max(0, (d2 - d1).days)


def _risk_label(vince_risk):
    """Classify risk from vince_risk dict into Low/Moderate/High risk."""
    if not vince_risk or not isinstance(vince_risk, dict):
        return ""
    # Prefer the volatility_label if already computed by the analyzer
    vol = vince_risk.get("volatility_label")
    if vol:
        vol_u = str(vol).upper()
        if "LOW" in vol_u:
            return "LOW RISK"
        if "MOD" in vol_u or "MEDIUM" in vol_u:
            return "MODERATE RISK"
        if "HIGH" in vol_u or "EXTREME" in vol_u:
            return "HIGH RISK"
    # Fallback: derive from optimal_f
    try:
        opt_f = vince_risk.get("optimal_f")
        if opt_f is None:
            return ""
        opt_f = float(opt_f)
        if opt_f <= 0.1:
            return "LOW RISK"
        elif opt_f <= 0.25:
            return "MODERATE RISK"
        else:
            return "HIGH RISK"
    except Exception:
        return ""


def _sizing_label(vince_risk, buy_price, quantity):
    """Return sizing classification (UNDERSIZED / OPTIMAL / OVERSIZED)."""
    if not vince_risk or not isinstance(vince_risk, dict):
        return ""
    # Prefer the precomputed label
    status = vince_risk.get("sizing_status")
    if status:
        return str(status).upper()
    # Fallback: compare current qty to recommended
    try:
        rec_qty = (
            vince_risk.get("recommended_shares")
            or vince_risk.get("recommended_quantity")
        )
        if rec_qty is None or not quantity:
            return ""
        rec_qty = float(rec_qty)
        cur = float(quantity)
        if rec_qty <= 0:
            return ""
        ratio = cur / rec_qty
        if ratio < 0.7:
            return "UNDERSIZED"
        elif ratio > 1.3:
            return "OVERSIZED"
        else:
            return "OPTIMAL"
    except Exception:
        return ""


def _build_open_export_rows():
    """Build rows for Open Positions export (includes live analysis data)."""
    positions = get_open_positions(user_id=_uid(), is_admin=_is_admin())
    if not positions:
        return []

    # Run analysis for live data
    try:
        analyses = analyze_all_open_positions(positions)
    except Exception:
        analyses = {}

    # Build lookup by position id.  analyze_all_open_positions returns a list of
    # dicts shaped like: { 'position': {id, ticker, ...}, 'holding': {...},
    # 'targets': {...}, 'recommendation': {...}, 'vince_risk': {...}, ... }
    analysis_map = {}
    if isinstance(analyses, list):
        for a in analyses:
            if not isinstance(a, dict):
                continue
            pos_obj = a.get("position") or {}
            pid = (
                pos_obj.get("id")
                if isinstance(pos_obj, dict)
                else None
            ) or a.get("position_id") or a.get("id")
            if pid is not None:
                analysis_map[pid] = a
    elif isinstance(analyses, dict):
        analysis_map = analyses

    rows = []
    for p in positions:
        pid = p.get("id")
        a = analysis_map.get(pid) or {}

        buy_price = float(p.get("buy_price") or 0)
        qty = int(p.get("quantity") or 0)
        invested = buy_price * qty

        # Extract current price from nested analysis structures
        holding = a.get("holding") or {}
        targets = a.get("targets") or {}
        indicators = a.get("indicators") or {}
        current = (
            holding.get("current_price")
            if isinstance(holding, dict) else None
        )
        if current is None and isinstance(targets, dict):
            current = targets.get("current_price")
        if current is None and isinstance(indicators, dict):
            current = indicators.get("price")
        if current is None:
            current = a.get("current_price") or a.get("latest_price")
        try:
            current = float(current) if current is not None else None
        except Exception:
            current = None

        # P&L — prefer analyzer values if present, else compute
        pnl = None
        pnl_pct = None
        if isinstance(holding, dict):
            pnl = (
                holding.get("pnl_amount")
                if holding.get("pnl_amount") is not None
                else (holding.get("pnl") if holding.get("pnl") is not None else holding.get("unrealized_pnl"))
            )
            pnl_pct = holding.get("pnl_pct")
        if pnl is None and isinstance(targets, dict):
            pnl = targets.get("pnl_amount")
            pnl_pct = pnl_pct if pnl_pct is not None else targets.get("pnl_pct")
        if pnl is None and current is not None and buy_price:
            pnl = (current - buy_price) * qty
            pnl_pct = ((current - buy_price) / buy_price) * 100.0
        try:
            pnl = float(pnl) if pnl is not None else None
        except Exception:
            pnl = None
        try:
            pnl_pct = float(pnl_pct) if pnl_pct is not None else None
        except Exception:
            pnl_pct = None

        days = _days_between(p.get("buy_date"))

        # Recommendation / Action
        rec_obj = a.get("recommendation")
        if isinstance(rec_obj, dict):
            recommendation = rec_obj.get("action") or rec_obj.get("verdict") or ""
        else:
            recommendation = (
                rec_obj
                or a.get("action")
                or (a.get("signal") or {}).get("action")
                or ""
            )

        vince_risk = a.get("vince_risk") or {}
        risk = _risk_label(vince_risk)
        sizing = _sizing_label(vince_risk, buy_price, qty)

        rows.append({
            "Ticker": (p.get("ticker") or "").replace(".NS", ""),
            "Strategy": p.get("strategy_code", ""),
            "Buy Price": buy_price,
            "Buy Date": p.get("buy_date", ""),
            "Quantity": qty,
            "Invested": round(invested, 2),
            "Current Price": round(current, 2) if current is not None else "",
            "P&L": round(pnl, 2) if pnl is not None else "",
            "P&L %": round(pnl_pct, 2) if pnl_pct is not None else "",
            "Days Held": days,
            "Action": str(recommendation).upper() if recommendation else "",
            "Risk": risk,
            "Sizing": sizing,
            "Notes": p.get("notes", "") or "",
        })
    return rows


def _build_closed_export_rows():
    """Build rows for Closed Positions export, including the same charges &
    tax breakdown the Trades dashboard computes — so the Excel/PDF reflects
    everything the user sees on /portfolio AND /trades for that position."""
    uid = _uid()
    is_admin = _is_admin()
    positions = get_closed_positions(user_id=uid, is_admin=is_admin)

    # Pre-build a {position_id: trade_id} map so we can surface the auto-logged
    # trade reference next to each closed position.
    try:
        all_trades = get_all_trades(user_id=uid, is_admin=is_admin)
        pos_to_trade = {
            int(t["position_id"]): t["id"]
            for t in all_trades
            if t.get("position_id")
        }
    except Exception:
        pos_to_trade = {}

    rows = []
    for p in positions:
        buy_price = float(p.get("buy_price") or 0)
        sell_price = float(p.get("sell_price") or 0)
        qty = int(p.get("quantity") or 0)
        gross_pnl = (sell_price - buy_price) * qty if (buy_price and sell_price) else 0
        pnl_pct = ((sell_price - buy_price) / buy_price * 100.0) if buy_price else 0
        days = _days_between(p.get("buy_date"), p.get("sell_date"))
        ticker_clean = (p.get("ticker") or "").replace(".NS", "")

        # Compute the same charges + tax block the Trades dashboard uses, so
        # users see one consistent number across Portfolio and Trades exports.
        charges_total = 0.0
        net_pnl = gross_pnl
        tax_amt = 0.0
        post_tax = gross_pnl
        tax_cat = ""
        try:
            if buy_price and sell_price and qty and p.get("buy_date") and p.get("sell_date"):
                pnl_obj = calculate_trade(
                    stock=ticker_clean, platform="zerodha",
                    trade_type="delivery", exchange="NSE",
                    quantity=qty, buy_price=buy_price, sell_price=sell_price,
                    buy_date=p.get("buy_date"), sell_date=p.get("sell_date"),
                ).to_dict()
                charges_total = pnl_obj.get("charges", {}).get("total", 0) or 0
                net_pnl = pnl_obj.get("net_pnl", net_pnl) or net_pnl
                tax_amt = pnl_obj.get("total_tax", 0) or 0
                post_tax = pnl_obj.get("post_tax_pnl", net_pnl) or net_pnl
                tax_cat = pnl_obj.get("tax_category", "") or ""
        except Exception:
            pass  # fall back to gross only — never let calc failure block export

        linked_tid = pos_to_trade.get(int(p["id"])) if p.get("id") else None

        rows.append({
            "Ticker":      ticker_clean,
            "Strategy":    p.get("strategy_code", ""),
            "Buy Price":   buy_price,
            "Buy Date":    p.get("buy_date", ""),
            "Sell Price":  sell_price,
            "Sell Date":   p.get("sell_date", "") or "",
            "Quantity":    qty,
            "Invested":    round(buy_price * qty, 2),
            "Sale Value":  round(sell_price * qty, 2),
            "Gross P&L":   round(gross_pnl, 2),
            "Charges":     round(charges_total, 2),
            "Net P&L":     round(net_pnl, 2),
            "Tax (incl. cess)": round(tax_amt, 2),
            "Post-Tax P&L": round(post_tax, 2),
            "Tax Category": tax_cat,
            "P&L %":       round(pnl_pct, 2),
            "Days Held":   days,
            "Reason":      p.get("sell_reason", "") or "",
            "Linked Trade #": linked_tid if linked_tid else "",
            "Notes":       p.get("notes", "") or "",
        })
    return rows


@app.route("/api/portfolio/export/xlsx")
def api_portfolio_export_xlsx():
    """Export portfolio positions to Excel (.xlsx)."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tab = (request.args.get("tab") or "open").lower()

    if tab == "closed":
        rows = _build_closed_export_rows()
        sheet_title = "Closed Positions"
        filename = f"portfolio_closed_{date.today().isoformat()}.xlsx"
    else:
        rows = _build_open_export_rows()
        sheet_title = "Open Positions"
        filename = f"portfolio_open_{date.today().isoformat()}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Title row
    title_text = f"Strategy Portfolio — {sheet_title}"
    ws.cell(row=1, column=1, value=title_text)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E78")

    # Generated-on row
    ws.cell(row=2, column=1, value=f"Generated: {date.today().strftime('%Y-%m-%d')} | Records: {len(rows)}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    if not rows:
        ws.cell(row=4, column=1, value="No positions to export.")
        # Merge title row across a few cols
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    else:
        headers = list(rows[0].keys())

        # Merge title across all columns
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

        # Header row (row 4)
        header_fill = PatternFill("solid", fgColor="2E5A8A")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin", color="888888")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data rows
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")
        green_fill = PatternFill("solid", fgColor="E6F4EA")
        red_fill = PatternFill("solid", fgColor="FDE7E9")

        money_cols = {
            "Buy Price", "Sell Price", "Current Price", "Invested",
            "Sale Value", "P&L", "Gross P&L", "Charges", "Net P&L",
            "Tax (incl. cess)", "Post-Tax P&L",
        }
        signed_cols = {"P&L", "P&L %", "Gross P&L", "Net P&L", "Post-Tax P&L"}

        for r_idx, row in enumerate(rows, start=5):
            for c_idx, h in enumerate(headers, start=1):
                val = row.get(h, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                # Color P&L-style columns by sign
                if h in signed_cols and isinstance(val, (int, float)):
                    if val > 0:
                        cell.font = green_font
                        cell.fill = green_fill
                    elif val < 0:
                        cell.font = red_font
                        cell.fill = red_fill
                # Format number columns
                if h in money_cols and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                elif h == "P&L %" and isinstance(val, (int, float)):
                    cell.number_format = '0.00"%"'

        # Auto column widths
        for c_idx, h in enumerate(headers, start=1):
            col_letter = get_column_letter(c_idx)
            max_len = len(str(h))
            for row in rows:
                v = row.get(h, "")
                l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Freeze top rows
        ws.freeze_panes = "A5"

        # Summary row at the bottom (for numeric columns)
        summary_row_idx = 5 + len(rows) + 1
        ws.cell(row=summary_row_idx, column=1, value="TOTAL").font = Font(bold=True)
        sum_cols = {
            "Quantity", "Invested", "Sale Value", "P&L", "Gross P&L",
            "Charges", "Net P&L", "Tax (incl. cess)", "Post-Tax P&L",
        }
        for c_idx, h in enumerate(headers, start=1):
            if h in sum_cols:
                col_sum = sum((row.get(h) or 0) for row in rows if isinstance(row.get(h), (int, float)))
                cell = ws.cell(row=summary_row_idx, column=c_idx, value=round(col_sum, 2))
                cell.font = Font(bold=True)
                if h != "Quantity":
                    cell.number_format = '#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/api/portfolio/export/pdf")
def api_portfolio_export_pdf():
    """Export portfolio positions to PDF."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    tab = (request.args.get("tab") or "open").lower()

    if tab == "closed":
        rows = _build_closed_export_rows()
        title = "Strategy Portfolio — Closed Positions"
        filename = f"portfolio_closed_{date.today().isoformat()}.pdf"
        # Closed has 20 columns (charges/tax breakdown) — needs landscape A3.
        page_size = landscape(A3)
    else:
        rows = _build_open_export_rows()
        title = "Strategy Portfolio — Open Positions"
        filename = f"portfolio_open_{date.today().isoformat()}.pdf"
        # Open has ~14 columns — landscape A4 is fine.
        page_size = landscape(A4)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PortfolioTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=colors.HexColor("#1F4E78"),
        alignment=0,
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "PortfolioSub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#666666"),
        spaceAfter=10,
    )
    # Header cells need to wrap onto multiple lines so 20 columns don't collide.
    th_style = ParagraphStyle(
        "PortfolioHeaderCell", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8, leading=10,
        textColor=colors.whitesmoke, alignment=1,  # center
    )
    td_left = ParagraphStyle(
        "PortfolioCellLeft", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7.5, leading=9.5,
        textColor=colors.HexColor("#1F2933"), alignment=0,
    )
    td_right = ParagraphStyle(
        "PortfolioCellRight", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7.5, leading=9.5,
        textColor=colors.HexColor("#1F2933"), alignment=2,  # right
    )

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(
        f"Generated: {date.today().strftime('%Y-%m-%d')} &nbsp;&nbsp; | &nbsp;&nbsp; "
        f"Records: {len(rows)}",
        sub_style
    ))

    if not rows:
        story.append(Paragraph("No positions to export.", styles["Normal"]))
    else:
        headers = list(rows[0].keys())

        # Money / signed / text column groups for formatting
        money_cols = {
            "Buy Price", "Sell Price", "Current Price", "Invested", "Sale Value",
            "P&L", "Gross P&L", "Charges", "Net P&L", "Tax (incl. cess)", "Post-Tax P&L",
        }
        right_cols = money_cols | {
            "Quantity", "P&L %", "Days Held", "Linked Trade #",
        }
        text_cols = {"Notes", "Reason", "Ticker", "Strategy", "Tax Category", "Action", "Risk", "Sizing"}

        def _fmt_cell(h, v):
            if v is None or v == "":
                return ""
            if isinstance(v, float):
                if h == "P&L %":
                    return f"{v:.2f}%"
                if h in money_cols:
                    return f"{v:,.2f}"
                return f"{v:.2f}"
            return str(v)

        # Build data with Paragraph cells so headers and long text wrap properly.
        header_row = [Paragraph(h.replace("&", "&amp;"), th_style) for h in headers]
        data = [header_row]

        for r in rows:
            row_vals = []
            for h in headers:
                v = r.get(h, "")
                txt = _fmt_cell(h, v)
                style = td_right if h in right_cols else td_left
                row_vals.append(Paragraph(txt.replace("&", "&amp;"), style))
            data.append(row_vals)

        # Compute column widths proportionally
        page_width = page_size[0] - 20 * mm

        # Width weights — longer-text columns get more, narrow numeric columns less.
        weight_map = {
            "Notes": 2.4, "Reason": 1.8,
            "Ticker": 1.0, "Strategy": 0.7,
            "Buy Date": 1.0, "Sell Date": 1.0,
            "Buy Price": 0.9, "Sell Price": 0.9, "Current Price": 0.9,
            "Quantity": 0.6, "Days Held": 0.6,
            "Invested": 1.0, "Sale Value": 1.0,
            "Gross P&L": 1.0, "Charges": 0.95, "Net P&L": 1.0,
            "Tax (incl. cess)": 1.1, "Post-Tax P&L": 1.05,
            "Tax Category": 0.85, "P&L": 1.0, "P&L %": 0.7,
            "Linked Trade #": 0.85, "Action": 1.1, "Risk": 0.9, "Sizing": 1.0,
        }
        weights = [weight_map.get(h, 1.0) for h in headers]
        total_w = sum(weights)
        col_widths = [page_width * (w / total_w) for w in weights]

        tbl = Table(data, colWidths=col_widths, repeatRows=1)

        ts = TableStyle([
            # Header band
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5A8A")),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),

            # Body
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 1), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AAAAAA")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.whitesmoke, colors.HexColor("#F3F6FA")]),
        ])

        # Re-apply colour to signed columns by replacing the Paragraph with a coloured one
        # (per-cell TEXTCOLOR doesn't apply to Paragraph children, so colour goes inline).
        signed_cols = ("P&L", "P&L %", "Gross P&L", "Net P&L", "Post-Tax P&L")
        amber_cols  = ("Charges",)
        purple_cols = ("Tax (incl. cess)",)

        def _coloured_paragraph(text: str, hex_color: str, align: int = 2) -> Paragraph:
            ps = ParagraphStyle(
                "Coloured", parent=td_right if align == 2 else td_left,
                textColor=colors.HexColor(hex_color),
                fontName="Helvetica-Bold",
            )
            return Paragraph(text, ps)

        for i, r in enumerate(rows, start=1):
            for h in signed_cols:
                if h not in headers:
                    continue
                v = r.get(h)
                if not isinstance(v, (int, float)) or v == 0:
                    continue
                idx = headers.index(h)
                colour = "#006100" if v > 0 else "#9C0006"
                data[i][idx] = _coloured_paragraph(_fmt_cell(h, v), colour, align=2)
            for h in amber_cols:
                if h in headers:
                    v = r.get(h)
                    if isinstance(v, (int, float)) and v:
                        data[i][headers.index(h)] = _coloured_paragraph(
                            _fmt_cell(h, v), "#B7791F", align=2)
            for h in purple_cols:
                if h in headers:
                    v = r.get(h)
                    if isinstance(v, (int, float)) and v:
                        data[i][headers.index(h)] = _coloured_paragraph(
                            _fmt_cell(h, v), "#6B2D9C", align=2)

        tbl.setStyle(ts)
        story.append(tbl)

        # Totals
        def _total(col):
            return sum((r.get(col) or 0) for r in rows if isinstance(r.get(col), (int, float)))

        total_invested = _total("Invested")
        total_pnl      = _total("P&L")
        total_qty      = _total("Quantity")
        total_gross    = _total("Gross P&L")
        total_charges  = _total("Charges")
        total_net      = _total("Net P&L")
        total_tax      = _total("Tax (incl. cess)")
        total_post     = _total("Post-Tax P&L")

        story.append(Spacer(1, 8))
        totals_style = ParagraphStyle(
            "Totals", parent=styles["Normal"], fontSize=10,
            textColor=colors.HexColor("#1F4E78"), leading=14,
        )
        bits = [
            f"<b>Total Positions:</b> {len(rows)}",
            f"<b>Total Quantity:</b> {int(total_qty)}",
        ]
        if total_invested:
            bits.append(f"<b>Total Invested:</b> Rs. {total_invested:,.2f}")
        if total_pnl:
            color = "#006100" if total_pnl >= 0 else "#9C0006"
            bits.append(f"<b>Total P&amp;L:</b> <font color='{color}'>Rs. {total_pnl:,.2f}</font>")
        if total_gross:
            color = "#006100" if total_gross >= 0 else "#9C0006"
            bits.append(f"<b>Gross P&amp;L:</b> <font color='{color}'>Rs. {total_gross:,.2f}</font>")
        if total_charges:
            bits.append(f"<b>Charges:</b> <font color='#B7791F'>Rs. {total_charges:,.2f}</font>")
        if total_net:
            color = "#006100" if total_net >= 0 else "#9C0006"
            bits.append(f"<b>Net P&amp;L:</b> <font color='{color}'>Rs. {total_net:,.2f}</font>")
        if total_tax:
            bits.append(f"<b>Tax:</b> <font color='#6B2D9C'>Rs. {total_tax:,.2f}</font>")
        if total_post:
            color = "#006100" if total_post >= 0 else "#9C0006"
            bits.append(f"<b>Post-Tax:</b> <font color='{color}'>Rs. {total_post:,.2f}</font>")
        story.append(Paragraph(" &nbsp;&nbsp;|&nbsp;&nbsp; ".join(bits), totals_style))

    doc.build(story)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  🚀  Bollinger Squeeze Web Dashboard")
    print(f"  📂  Data dir: {CSV_DIR}")
    print(f"  🌐  Open: http://127.0.0.1:5001\n")
    app.run(debug=False, port=5001)
