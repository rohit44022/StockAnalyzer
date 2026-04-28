"""
excel_top_picks.py — Excel exporter for the Top 5 Picks dashboard.

Produces a colour-coded multi-sheet workbook from a `find_top_picks()` result:

  ┌────────────────────────────────────────────────────────────┐
  │ Sheet 1 — Summary                                          │
  │   • Header band (BUY = green, SELL = red)                  │
  │   • Run metadata + scan funnel                             │
  │   • One row per top pick with composite score, grade,      │
  │     verdicts, R:R, target — colour-coded                   │
  ├────────────────────────────────────────────────────────────┤
  │ Sheet 2..N — Detail per pick                               │
  │   • Header banner with rank/ticker/composite/grade          │
  │   • 7 component score bars (data-bar style)                 │
  │   • BB + TA + Triple verdict block                          │
  │   • Triple-Conviction breakdown (BB/TA/PA totals)           │
  │   • Unified Triple Targets (entry, stop, T1/T2/T3)          │
  │   • Reasons (green) + Warnings (amber)                      │
  │   • Top action items                                        │
  │   • Data quality footer                                     │
  └────────────────────────────────────────────────────────────┘

The module is purely additive — calling it does not touch any existing data
shape or behaviour.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Color, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ─────────────────────────── colour palette ───────────────────────────
# Lighter than dashboard hex so text stays readable on white background.

WHITE         = "FFFFFF"
BORDER_GREY   = "B0B7BD"
HEADER_BLUE   = "1F4E78"
HEADER_GREEN  = "1E7E34"
HEADER_RED    = "B83227"
HEADER_AMBER  = "B8860B"
SUB_GREY      = "5C6770"
TEXT_DARK     = "1F2933"

GREEN_DARK    = "0E6B2E"
GREEN_LIGHT   = "D8F0DD"
RED_DARK      = "9C0C0C"
RED_LIGHT     = "FBE0DE"
AMBER_DARK    = "8B6A0E"
AMBER_LIGHT   = "FFF1C8"
BLUE_DARK     = "0E4F88"
BLUE_LIGHT    = "DCEAF7"
PURPLE_DARK   = "5A2D85"
PURPLE_LIGHT  = "ECDFFA"

# Bar fills used to draw the "score bar" effect in the detail sheet
BAR_HIGH      = "5CC36F"   # 70+
BAR_MED       = "5BA8E8"   # 50–69
BAR_LOW       = "E0B547"   # 35–49
BAR_BAD       = "E26757"   # <35
BAR_TRACK     = "E5E7EB"


# ─────────────────────────── helpers ───────────────────────────

def _thin_border() -> Border:
    s = Side(border_style="thin", color=BORDER_GREY)
    return Border(left=s, right=s, top=s, bottom=s)


def _no_border() -> Border:
    return Border()


def _safe(v: Any, default: Any = "") -> Any:
    """Return v if usable, else default. Treats None/NaN as missing."""
    if v is None:
        return default
    if isinstance(v, float):
        try:
            import math
            if math.isnan(v) or math.isinf(v):
                return default
        except Exception:
            pass
    return v


def _score_color(score: float) -> tuple[str, str]:
    """(font_color_hex, fill_hex) for a 0-100 score."""
    if score is None:
        return TEXT_DARK, "F0F2F5"
    try:
        s = float(score)
    except Exception:
        return TEXT_DARK, "F0F2F5"
    if s >= 80: return GREEN_DARK, GREEN_LIGHT
    if s >= 60: return BLUE_DARK,  BLUE_LIGHT
    if s >= 45: return AMBER_DARK, AMBER_LIGHT
    return RED_DARK, RED_LIGHT


def _bar_color(score: float) -> str:
    """Just the bar fill colour (matches dashboard palette)."""
    try:
        s = float(score)
    except Exception:
        return BAR_TRACK
    if s >= 70: return BAR_HIGH
    if s >= 50: return BAR_MED
    if s >= 35: return BAR_LOW
    return BAR_BAD


def _grade_color(grade: str) -> tuple[str, str]:
    g = (grade or "").upper().strip()
    if g in ("A+", "A"):    return GREEN_DARK, GREEN_LIGHT
    if g in ("B+", "B"):    return BLUE_DARK,  BLUE_LIGHT
    if g == "C":            return AMBER_DARK, AMBER_LIGHT
    if g in ("D", "F"):     return RED_DARK,   RED_LIGHT
    return TEXT_DARK, "F0F2F5"


def _verdict_color(verdict: str) -> tuple[str, str]:
    v = (verdict or "").upper()
    if "STRONG SELL" in v or "SUPER STRONG SELL" in v:
        return RED_DARK, RED_LIGHT
    if "SELL" in v:
        return RED_DARK, RED_LIGHT
    if "STRONG BUY" in v or "SUPER STRONG BUY" in v:
        return GREEN_DARK, GREEN_LIGHT
    if "BUY" in v:
        return GREEN_DARK, GREEN_LIGHT
    if "HOLD" in v or "WAIT" in v or "NEUTRAL" in v:
        return AMBER_DARK, AMBER_LIGHT
    return SUB_GREY, "F0F2F5"


def _direction_color(signal_filter: str) -> tuple[str, str]:
    if (signal_filter or "").upper() == "SELL":
        return WHITE, HEADER_RED
    return WHITE, HEADER_GREEN


def _set_cell(ws: Worksheet, row: int, col: int, value: Any,
              *, bold: bool = False, italic: bool = False,
              size: float = 10, color: str = TEXT_DARK,
              fill: Optional[str] = None,
              halign: str = "left", valign: str = "center",
              wrap: bool = False, border: bool = True,
              number_format: Optional[str] = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = _thin_border() if border else _no_border()
    if number_format:
        cell.number_format = number_format


def _autosize(ws: Worksheet, widths: dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _ticker_display(t: str) -> str:
    return (t or "").replace(".NS", "").replace(".BO", "")


def _safe_sheet_name(name: str) -> str:
    """Excel disallows []:*?/\\ and 31-char max."""
    bad = "[]:*?/\\"
    out = "".join(c for c in (name or "") if c not in bad)
    return out[:31] or "Sheet"


# ─────────────────────────── Summary sheet ───────────────────────────

def _build_summary_sheet(ws: Worksheet, picks: list[dict],
                         method: str, signal_filter: str,
                         total_scanned: int, total_signals: int,
                         total_qualified: int, total_analyzed: int) -> None:
    ws.title = "Summary"

    # ── Title band ──
    head_font_color, head_fill = _direction_color(signal_filter)
    title = f"Top {len(picks)} {signal_filter.upper()} Picks — Method {method}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    _set_cell(ws, 1, 1, title, bold=True, size=16,
              color=head_font_color, fill=head_fill,
              halign="center", border=False)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    _set_cell(ws, 2, 1,
              f"Generated {date.today().strftime('%d %b %Y')} · "
              f"Hiranya Triple Conviction Engine",
              italic=True, size=9, color=SUB_GREY,
              halign="center", border=False)
    ws.row_dimensions[2].height = 18

    # ── Funnel band ──
    funnel = [
        ("Total scanned",  total_scanned),
        ("With signals",   total_signals),
        ("Qualified",      total_qualified),
        ("Deep analyzed",  total_analyzed),
        ("Top picks",      len(picks)),
    ]
    for i, (lbl, val) in enumerate(funnel, start=1):
        col = (i - 1) * 2 + 1
        _set_cell(ws, 4, col,     lbl, bold=True, size=9,
                  color=SUB_GREY, fill="F0F2F5", halign="center")
        _set_cell(ws, 4, col + 1, val, size=11, color=TEXT_DARK,
                  fill="F8FAFC", halign="center", number_format="#,##0")

    # ── Header row ──
    headers = [
        "Rank", "Ticker", "Price", "Composite", "Grade",
        "BB", "TA", "Triple", "Conf %", "R:R",
        "Target", "Stop", "Reasons", "Warnings",
    ]
    for c, h in enumerate(headers, start=1):
        _set_cell(ws, 6, c, h, bold=True, size=10, color=WHITE,
                  fill=HEADER_BLUE, halign="center")
    ws.row_dimensions[6].height = 22

    # ── Pick rows ──
    rank_emoji = {1: "🥇", 2: "🥈", 3: "🥉"}
    for i, p in enumerate(picks, start=1):
        row = 6 + i
        rank      = p.get("rank", i)
        ticker    = _ticker_display(p.get("ticker"))
        price     = _safe(p.get("current_price"))
        composite = _safe(p.get("composite_score"))
        grade     = _safe(p.get("grade"))
        bb_v      = _safe(p.get("bb_signal_type"))
        ta_v      = _safe(p.get("ta_verdict"))
        triple_v  = _safe(p.get("triple_verdict"))
        conf      = _safe(p.get("triple_confidence"))
        rr        = _safe(p.get("rr_ratio"))
        target    = _safe(p.get("target_upside")) or _safe(p.get("target_downside"))
        stop      = _safe(p.get("stop_loss"))
        reasons   = "; ".join((p.get("reasons") or [])[:3])
        warnings  = "; ".join((p.get("warnings") or [])[:3])

        comp_color, comp_fill = _score_color(composite if isinstance(composite, (int, float)) else 0)
        grade_color, grade_fill = _grade_color(grade if isinstance(grade, str) else "")
        bb_color,  bb_fill  = _verdict_color(bb_v if isinstance(bb_v, str) else "")
        ta_color,  ta_fill  = _verdict_color(ta_v if isinstance(ta_v, str) else "")
        tv_color,  tv_fill  = _verdict_color(triple_v if isinstance(triple_v, str) else "")

        _set_cell(ws, row, 1, f"{rank_emoji.get(rank, str(rank))}",
                  bold=True, halign="center")
        _set_cell(ws, row, 2, ticker, bold=True, halign="center")
        _set_cell(ws, row, 3, price, halign="right", number_format='"₹"#,##0.00')
        _set_cell(ws, row, 4, composite, bold=True, color=comp_color, fill=comp_fill,
                  halign="center", number_format="0.0")
        _set_cell(ws, row, 5, grade, bold=True, color=grade_color, fill=grade_fill,
                  halign="center")
        _set_cell(ws, row, 6, bb_v,     color=bb_color, fill=bb_fill, halign="center")
        _set_cell(ws, row, 7, ta_v,     color=ta_color, fill=ta_fill, halign="center")
        _set_cell(ws, row, 8, triple_v, color=tv_color, fill=tv_fill, halign="center")
        _set_cell(ws, row, 9, conf, halign="center", number_format='0.0"%"')
        _set_cell(ws, row, 10, rr, halign="center",
                  number_format='"1:"0.00' if isinstance(rr, (int, float)) else "@")
        _set_cell(ws, row, 11, target, halign="right",
                  number_format='"₹"#,##0.00' if isinstance(target, (int, float)) else "@")
        _set_cell(ws, row, 12, stop, halign="right",
                  number_format='"₹"#,##0.00' if isinstance(stop, (int, float)) else "@")
        _set_cell(ws, row, 13, reasons, wrap=True, color=GREEN_DARK)
        _set_cell(ws, row, 14, warnings, wrap=True, color=AMBER_DARK)
        ws.row_dimensions[row].height = 32

    # ── Footer note ──
    foot_row = 6 + max(len(picks), 1) + 2
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=14)
    _set_cell(ws, foot_row, 1,
              "Open the per-pick tabs (one for each ticker) for the full breakdown — "
              "components, BB/TA/Triple, target & stop, reasons, warnings, action items.",
              italic=True, size=9, color=SUB_GREY, halign="center", border=False)

    _autosize(ws, {
        1: 6, 2: 12, 3: 12, 4: 11, 5: 8, 6: 14, 7: 14, 8: 14,
        9: 9, 10: 9, 11: 12, 12: 12, 13: 50, 14: 50,
    })


# ─────────────────────────── Detail sheet ───────────────────────────

_COMP_ORDER = [
    ("bb_strategy",      "BB Strategy"),
    ("ta_score",         "TA Score"),
    ("triple_score",     "Triple Score"),
    ("pa_score",         "Price Action"),
    ("risk_reward",      "Risk/Reward"),
    ("signal_agreement", "Agreement"),
    ("data_quality",     "Data Quality"),
]


def _draw_score_bar(ws: Worksheet, row: int, start_col: int,
                    score: float, width_cells: int = 10) -> None:
    """Render a fake score bar across `width_cells` cells.
    Filled cells get the score colour; remaining cells get the track colour.
    """
    try:
        pct = max(0.0, min(1.0, float(score) / 100.0))
    except Exception:
        pct = 0.0
    filled = int(round(pct * width_cells))
    fill_color = _bar_color(score)
    for k in range(width_cells):
        c = start_col + k
        _set_cell(ws, row, c, "", border=True,
                  fill=fill_color if k < filled else BAR_TRACK)


def _build_detail_sheet(ws: Worksheet, p: dict, signal_filter: str) -> None:
    rank   = p.get("rank", "?")
    ticker = _ticker_display(p.get("ticker"))
    ws.title = _safe_sheet_name(f"#{rank} {ticker}")

    # ── Title banner ──
    head_color, head_fill = _direction_color(signal_filter)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    title = (
        f"#{rank}  ·  {ticker}  ·  ₹{(_safe(p.get('current_price'),0) or 0):,.2f}  "
        f"·  Composite {(_safe(p.get('composite_score'),0) or 0):.1f}/100  ·  "
        f"Grade {p.get('grade','—')}"
    )
    _set_cell(ws, 1, 1, title, bold=True, size=14,
              color=head_color, fill=head_fill, halign="center", border=False)
    ws.row_dimensions[1].height = 28

    # Sub-banner — direction, scan source, generated date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    _set_cell(ws, 2, 1,
              f"{signal_filter} pick  ·  {p.get('bb_signal_type','—')} "
              f"({(_safe(p.get('bb_confidence'),0) or 0):.0f}% BB confidence)  ·  "
              f"Generated {date.today().strftime('%d %b %Y')}",
              italic=True, size=9, color=SUB_GREY, halign="center", border=False)

    row = 4

    # ── Section: Component scores with bars ──
    _section_header(ws, row, "Composite Score Components", span=14)
    row += 1
    _set_cell(ws, row, 1, "Component", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 2, "Score",     bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 3, "Weight",    bold=True, fill="F0F2F5", halign="center")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=13)
    _set_cell(ws, row, 4, "Bar",       bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 14, "Weighted", bold=True, fill="F0F2F5", halign="center")
    row += 1

    comps = p.get("components") or {}
    for key, label in _COMP_ORDER:
        c = comps.get(key) or {}
        score    = _safe(c.get("score"), 0)
        weight   = _safe(c.get("weight"), 0)
        weighted = _safe(c.get("weighted"), 0)
        sc_color, sc_fill = _score_color(score)
        _set_cell(ws, row, 1, label, bold=True)
        _set_cell(ws, row, 2, score, color=sc_color, fill=sc_fill,
                  halign="center", number_format="0.0")
        _set_cell(ws, row, 3, weight, halign="center",
                  number_format="0.0%")
        # 10-cell wide bar across cols 4..13
        _draw_score_bar(ws, row, 4, score, width_cells=10)
        _set_cell(ws, row, 14, weighted, halign="center", number_format="0.0")
        row += 1

    # Hint row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    _set_cell(ws, row, 1,
              "Bars: green = 70+, blue = 50-69, amber = 35-49, red < 35.  "
              "Weighted column shows the contribution of each component to the composite.",
              italic=True, size=9, color=SUB_GREY, halign="left", border=False)
    row += 2

    # ── Section: System verdicts ──
    _section_header(ws, row, "System Verdicts", span=14)
    row += 1
    headers = ["Engine", "Verdict", "Score / Detail"]
    for c, h in enumerate(headers, start=1):
        col_span = 1 if c < 3 else 12
        if c == 3:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=14)
        _set_cell(ws, row, c if c < 3 else 3, h, bold=True,
                  fill="F0F2F5", halign="center")
    row += 1

    rows_verdict = [
        ("BB Strategy",
         p.get("bb_signal_type"),
         f"{p.get('bb_phase','')} · squeeze {'ON' if p.get('bb_squeeze_on') else 'OFF'} ({p.get('bb_squeeze_days',0)}d)"),
        ("Technical Analysis",
         p.get("ta_verdict"),
         f"raw {(_safe(p.get('ta_score'),0) or 0):+.0f}/100"),
        ("Triple Conviction",
         p.get("triple_verdict"),
         f"{p.get('triple_emoji','')} combined {(_safe(p.get('triple_combined_score'),0) or 0):+.0f}/{p.get('triple_max_score',425)} · "
         f"alignment {p.get('triple_alignment','')}"),
        ("Price Action",
         (p.get("pa_data") or {}).get("pa_verdict") if p.get("pa_data") else "—",
         f"setup: {(p.get('pa_data') or {}).get('setup_type','—') if p.get('pa_data') else '—'}"),
    ]
    for engine, verdict, detail in rows_verdict:
        v_color, v_fill = _verdict_color(verdict if isinstance(verdict, str) else "")
        _set_cell(ws, row, 1, engine, bold=True)
        _set_cell(ws, row, 2, verdict or "—", bold=True,
                  color=v_color, fill=v_fill, halign="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=14)
        _set_cell(ws, row, 3, detail, italic=True, color=SUB_GREY, halign="left")
        row += 1
    row += 1

    # ── Section: Triple Conviction subsystem totals ──
    _section_header(ws, row, "Triple Conviction Subsystem Scores", span=14)
    row += 1
    sub_scores = [
        ("BB Total (from Triple Engine)",  p.get("triple_bb_total"), 100),
        ("TA Total (from Triple Engine)",  p.get("triple_ta_total"), 100),
        ("PA Total (from Triple Engine)",  p.get("triple_pa_total"), 100),
        ("Cross-System Agreement",         p.get("triple_agreement"), 100),
        ("Triple Confidence",              p.get("triple_confidence"), 100),
    ]
    _set_cell(ws, row, 1, "Metric", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 2, "Value",  bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 3, "Of",     bold=True, fill="F0F2F5", halign="center")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
    _set_cell(ws, row, 4, "Visual", bold=True, fill="F0F2F5", halign="center")
    row += 1
    for label, val, mx in sub_scores:
        v = _safe(val, 0) or 0
        # Map -100..100 → 0..100 for the bar (only for the BB/TA/PA totals which are signed)
        if label.startswith(("BB Total", "TA Total", "PA Total")):
            visual = max(0, min(100, (v + 100) / 2))
        else:
            visual = max(0, min(100, v))
        sc_color, sc_fill = _score_color(visual)
        _set_cell(ws, row, 1, label, bold=True)
        _set_cell(ws, row, 2, v, color=sc_color, fill=sc_fill,
                  halign="center", number_format="0.0")
        _set_cell(ws, row, 3, mx, halign="center")
        _draw_score_bar(ws, row, 4, visual, width_cells=11)
        row += 1
    row += 1

    # ── Section: Unified Triple Targets / Trade Plan ──
    _section_header(ws, row, "Trade Plan — Unified Triple Targets", span=14)
    row += 1
    tt = p.get("triple_targets") or {}
    direction = tt.get("direction") or signal_filter
    entry = (tt.get("entry") or {}).get("price") or _safe(p.get("current_price"))
    zone  = (tt.get("entry") or {}).get("zone") or [None, None]
    stop  = (tt.get("stop_loss") or {}).get("price") or _safe(p.get("stop_loss"))
    stop_pct = (tt.get("stop_loss") or {}).get("pct")
    targets = tt.get("targets") or []

    # Plan summary row
    _set_cell(ws, row, 1, "Direction", bold=True, fill="F0F2F5", halign="center")
    dir_color, dir_fill = _verdict_color(direction)
    _set_cell(ws, row, 2, direction or "—", bold=True,
              color=dir_color, fill=dir_fill, halign="center")
    _set_cell(ws, row, 3, "Entry", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 4, entry, halign="right",
              number_format='"₹"#,##0.00' if isinstance(entry, (int, float)) else "@")
    _set_cell(ws, row, 5, "Zone", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 6,
              f"₹{zone[0]:.2f} – ₹{zone[1]:.2f}" if zone[0] and zone[1] else "—",
              halign="center")
    _set_cell(ws, row, 7, "Stop", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 8, stop, color=RED_DARK, fill=RED_LIGHT, halign="right",
              number_format='"₹"#,##0.00' if isinstance(stop, (int, float)) else "@")
    _set_cell(ws, row, 9, "Stop %", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 10, stop_pct, color=RED_DARK, halign="center",
              number_format='+0.00"%";-0.00"%"' if isinstance(stop_pct, (int, float)) else "@")
    _set_cell(ws, row, 11, "R:R", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 12, _safe(p.get("rr_ratio")), halign="center",
              number_format='"1:"0.00' if isinstance(p.get("rr_ratio"), (int, float)) else "@")
    _set_cell(ws, row, 13, "Conf %", bold=True, fill="F0F2F5", halign="center")
    _set_cell(ws, row, 14, _safe(p.get("triple_confidence")), halign="center",
              number_format='0.0"%"' if isinstance(p.get("triple_confidence"), (int, float)) else "@")
    row += 1

    # Targets table
    if targets:
        row += 1
        _set_cell(ws, row, 1, "Tier",     bold=True, fill="F0F2F5", halign="center")
        _set_cell(ws, row, 2, "Price",    bold=True, fill="F0F2F5", halign="center")
        _set_cell(ws, row, 3, "% Move",   bold=True, fill="F0F2F5", halign="center")
        _set_cell(ws, row, 4, "R:R",      bold=True, fill="F0F2F5", halign="center")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=14)
        _set_cell(ws, row, 5, "Sources",  bold=True, fill="F0F2F5", halign="center")
        row += 1
        for t in targets[:3]:
            tier_label = t.get("tier", "—")
            tprice  = _safe(t.get("price"))
            tpct    = _safe(t.get("pct"))
            trr     = _safe(t.get("rr"))
            srcs    = ", ".join(t.get("sources") or [])
            up = isinstance(tpct, (int, float)) and tpct >= 0
            _set_cell(ws, row, 1, tier_label, bold=True,
                      color=GREEN_DARK if up else RED_DARK,
                      fill=GREEN_LIGHT if up else RED_LIGHT, halign="center")
            _set_cell(ws, row, 2, tprice, halign="right",
                      number_format='"₹"#,##0.00' if isinstance(tprice, (int, float)) else "@")
            _set_cell(ws, row, 3, tpct, halign="center",
                      number_format='+0.00"%";-0.00"%"' if isinstance(tpct, (int, float)) else "@")
            _set_cell(ws, row, 4, trr, halign="center",
                      number_format='"1:"0.00' if isinstance(trr, (int, float)) else "@")
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=14)
            _set_cell(ws, row, 5, srcs or "—", italic=True, color=SUB_GREY, halign="left")
            row += 1
    row += 1

    # ── Section: Reasons ──
    reasons = p.get("reasons") or []
    if reasons:
        _section_header(ws, row, f"Reasons ({len(reasons)})",
                        span=14, fill_color=GREEN_DARK)
        row += 1
        for r in reasons:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
            _set_cell(ws, row, 1, f"✓  {r}", color=GREEN_DARK, fill=GREEN_LIGHT,
                      wrap=True, halign="left")
            row += 1
        row += 1

    # ── Section: Warnings ──
    warnings = p.get("warnings") or []
    if warnings:
        _section_header(ws, row, f"Warnings ({len(warnings)})",
                        span=14, fill_color=HEADER_AMBER)
        row += 1
        for w in warnings:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
            _set_cell(ws, row, 1, f"⚠  {w}", color=AMBER_DARK, fill=AMBER_LIGHT,
                      wrap=True, halign="left")
            row += 1
        row += 1

    # ── Section: Action items ──
    actions = p.get("ta_action_items") or []
    if actions:
        _section_header(ws, row, "Top TA Action Items", span=14)
        row += 1
        for a in actions[:6]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
            _set_cell(ws, row, 1, f"→  {a}", color=BLUE_DARK, fill=BLUE_LIGHT,
                      wrap=True, halign="left")
            row += 1
        row += 1

    # ── Section: TA category breakdown ──
    cats = p.get("ta_categories") or {}
    if cats:
        _section_header(ws, row, "TA Category Breakdown (Murphy 6)", span=14)
        row += 1
        _set_cell(ws, row, 1, "Category", bold=True, fill="F0F2F5", halign="center")
        _set_cell(ws, row, 2, "Score",    bold=True, fill="F0F2F5", halign="center")
        _set_cell(ws, row, 3, "Max",      bold=True, fill="F0F2F5", halign="center")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
        _set_cell(ws, row, 4, "Visual",   bold=True, fill="F0F2F5", halign="center")
        row += 1
        for cat_name, cat_data in cats.items():
            score = _safe((cat_data or {}).get("score"), 0)
            mx    = _safe((cat_data or {}).get("max"), 0) or 1
            try:
                pct = abs(float(score)) / float(mx) * 100
            except Exception:
                pct = 0
            sc_color, sc_fill = _score_color(pct)
            _set_cell(ws, row, 1, cat_name.replace("_", " ").title(), bold=True)
            _set_cell(ws, row, 2, score, color=sc_color, fill=sc_fill,
                      halign="center", number_format="+0.0;-0.0")
            _set_cell(ws, row, 3, mx, halign="center", number_format="0")
            _draw_score_bar(ws, row, 4, pct, width_cells=11)
            row += 1
        row += 1

    # ── Section: Data quality footer ──
    _section_header(ws, row, "Data Quality", span=14)
    row += 1
    df_info = p.get("data_freshness") or {}
    last_date = df_info.get("last_date") or "—"
    stale = df_info.get("trading_days_stale", 0)
    fresh_color = GREEN_DARK if stale <= 1 else AMBER_DARK if stale <= 5 else RED_DARK
    fresh_fill  = GREEN_LIGHT if stale <= 1 else AMBER_LIGHT if stale <= 5 else RED_LIGHT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    _set_cell(ws, row, 1,
              f"Last bar: {last_date}  ·  {stale} trading day(s) stale",
              color=fresh_color, fill=fresh_fill, halign="center")
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    _set_cell(ws, row, 1,
              "Hiranya Top Picks Engine · For informational use only — not financial advice",
              italic=True, size=8, color=SUB_GREY, halign="center", border=False)

    # Column widths — give the bar columns some space, action/reason cols room to wrap
    _autosize(ws, {
        1: 22, 2: 14, 3: 10, 4: 6, 5: 6, 6: 6, 7: 6, 8: 8, 9: 8,
        10: 8, 11: 8, 12: 9, 13: 9, 14: 11,
    })


def _section_header(ws: Worksheet, row: int, text: str, *,
                    span: int = 14, fill_color: str = HEADER_BLUE) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    _set_cell(ws, row, 1, text, bold=True, size=11,
              color=WHITE, fill=fill_color, halign="left", border=False)
    ws.row_dimensions[row].height = 22


# ─────────────────────────── public entry ───────────────────────────

def build_top_picks_xlsx(result: dict, method: str, signal_filter: str) -> bytes:
    """Render the Top Picks result dict into a coloured xlsx workbook.

    `result` is the dict returned by `top_picks.engine.find_top_picks()`.
    Returns raw .xlsx bytes ready to stream as a download.
    """
    result = result or {}
    picks = result.get("picks") or []

    wb = Workbook()
    summary_ws = wb.active

    _build_summary_sheet(
        summary_ws, picks, method, signal_filter,
        total_scanned=int(result.get("total_scanned") or 0),
        total_signals=int(result.get("total_signals") or 0),
        total_qualified=int(result.get("total_qualified") or 0),
        total_analyzed=int(result.get("total_analyzed") or 0),
    )

    # If no picks, drop a friendly message on a single empty detail tab
    if not picks:
        ws = wb.create_sheet(title="No picks")
        _set_cell(ws, 1, 1,
                  result.get("message") or
                  "No stocks qualified under the strict picker criteria.",
                  italic=True, color=SUB_GREY, border=False)
    else:
        for p in picks:
            ws = wb.create_sheet(title=_safe_sheet_name(
                f"#{p.get('rank','?')} {_ticker_display(p.get('ticker'))}"
            ))
            _build_detail_sheet(ws, p, signal_filter)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
