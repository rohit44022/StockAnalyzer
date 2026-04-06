"""
exporter.py — Excel export engine for the Bollinger Band Squeeze Strategy Analyser.

Exports ALL scan/analysis results to a fully colour-coded, annotated Excel workbook.
Each cell has explanatory notes.  Every sheet has a dedicated LEGEND tab.

Sheet layout
────────────
  Sheet 1 : RESULTS          – main data table (colour-coded rows + per-cell explanations)
  Sheet 2 : LEGEND           – full glossary of every column, colour code, and indicator
  Sheet 3 : PHASE_GUIDE      – detailed explanation of the 3 squeeze phases
  Sheet 4 : HOW_TO_READ      – step-by-step reading guide for beginners
"""

import os
from datetime import datetime
from typing import Optional

# ── openpyxl imports ──────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bb_squeeze.signals import SignalResult
from bb_squeeze.fundamentals import FundamentalData


# ══════════════════════════════════════════════════════════════════
#  COLOUR PALETTE  (hex without #)
# ══════════════════════════════════════════════════════════════════

# ── Row background colours (signal type) ──────────────────────────
ROW_BUY        = "C6EFCE"   # pale green
ROW_SELL       = "FFC7CE"   # pale red
ROW_HOLD       = "BDD7EE"   # pale blue
ROW_WAIT       = "FFEB9C"   # pale yellow / amber
ROW_HEAD_FAKE  = "FFD966"   # amber / orange
ROW_MONITOR    = "F2F2F2"   # light grey
ROW_ERROR      = "EDEDED"   # grey

# ── Header / title fills ──────────────────────────────────────────
HDR_DARK       = "1F3864"   # dark navy – main header
HDR_GROUP_A    = "2E4057"   # Group A header
HDR_GROUP_B    = "21618C"   # Group B header
HDR_GROUP_C    = "145A32"   # Group C header
HDR_FUND       = "4A235A"   # Fundamentals header
HDR_META       = "1A5276"   # Meta / info header

# ── Font colours ──────────────────────────────────────────────────
WHITE          = "FFFFFF"
BLACK          = "000000"
DARK_GREEN     = "1D6A1D"
DARK_RED       = "9C0006"
DARK_BLUE      = "1F3864"
ORANGE         = "C65911"

# ── Sub-cell indicator fills ──────────────────────────────────────
GREEN_CELL     = "92D050"   # strong positive value
RED_CELL       = "FF0000"   # strong negative value  (used for MFI < 20, etc.)
AMBER_CELL     = "FFC000"   # caution
BLUE_CELL      = "4472C4"   # neutral / informational

# ── Legend sheet ─────────────────────────────────────────────────
LEG_HEADER     = "2C3E50"
LEG_SUBHEAD    = "2980B9"
LEG_ROW_ALT    = "EAF4FC"

THIN_BORDER_SIDE  = Side(style="thin", color="BFBFBF")
THIN_BORDER       = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
                           top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE)
MED_BORDER_SIDE   = Side(style="medium", color="595959")
MED_BORDER        = Border(left=MED_BORDER_SIDE, right=MED_BORDER_SIDE,
                           top=MED_BORDER_SIDE,  bottom=MED_BORDER_SIDE)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=BLACK, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_col_width(ws: Worksheet, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════════════════
#  COLUMN DEFINITIONS  (label, width, explanation comment)
# ══════════════════════════════════════════════════════════════════

# Each entry: (excel_header, col_width, tooltip / explanation text)
COLUMNS = [
    # ── Identity ──────────────────────────────────────────────────
    ("Ticker",           14, "NSE ticker symbol (e.g. RELIANCE.NS). Always ends in .NS for National Stock Exchange stocks."),
    ("Company Name",     26, "Full company name fetched from Yahoo Finance. Shown as N/A if fundamentals not fetched."),
    ("Sector",           18, "Business sector (e.g. Energy, IT, Banking). Helps compare stocks within same sector."),
    ("Scan Date",        14, "Date and time this scan/analysis was run. Data is based on closing prices of the previous trading day."),
    ("Last Data Date",   14, "The most recent trading date available in the local CSV file for this stock."),

    # ── Signal ────────────────────────────────────────────────────
    ("SIGNAL",           16, "The primary action signal.\n"
                             "BUY    = All 5 conditions met. Enter at tomorrow's 9:15 AM NSE open.\n"
                             "HOLD   = You are in a trade. Trend intact. Stay in.\n"
                             "SELL   = One of 3 exit signals triggered. Exit at tomorrow's open.\n"
                             "WAIT   = Squeeze is set but breakout not confirmed yet. Watch daily.\n"
                             "HEAD FAKE = Price broke out but indicators contradict it. DO NOT ENTER.\n"
                             "MONITOR = No setup. Keep on watchlist."),
    ("Confidence /100",  16, "Signal confidence score from 0 to 100.\n"
                             "Scoring:\n"
                             "  +25 pts  Condition 1: Squeeze ON (BBW at 6-month low)\n"
                             "  +25 pts  Condition 2: Price above upper Bollinger Band\n"
                             "  +20 pts  Condition 3: Volume green & above 50-day SMA\n"
                             "  +15 pts  Condition 4: CMF positive (smart money flowing in)\n"
                             "  +15 pts  Condition 5: MFI above 50 and rising\n"
                             "  + 5 pts  Bonus: CMF > +0.10 (strong accumulation)\n"
                             "  + 5 pts  Bonus: MFI > 80 (maximum fuel)\n"
                             "  Total max = 100\n\n"
                             "Position sizing guide:\n"
                             "  85-100 → Full position\n"
                             "  70-84  → 75% position\n"
                             "  60-69  → 50% position\n"
                             "  <60    → Consider skipping"),
    ("Action Message",   50, "Plain-English instruction explaining exactly what to do and why. "
                             "Includes stop loss level for BUY signals, exit reasons for SELL signals, "
                             "and direction hints for WAIT signals."),

    # ── Phase ─────────────────────────────────────────────────────
    ("Phase",            18, "Current squeeze phase. Bollinger identified 3 phases:\n\n"
                             "Phase 1 — COMPRESSION\n"
                             "  BBW is at its 6-month lowest value. Bands are maximally tight.\n"
                             "  Volatility has collapsed. The stock is 'coiling the spring'.\n"
                             "  CMF ≈ 0 (no clear flow). MFI ≈ 50 (balanced).\n"
                             "  ACTION: Add to watchlist. Set alert above upper band.\n\n"
                             "Phase 2 — DIRECTION CLUES\n"
                             "  Still in squeeze but CMF and MFI are beginning to show direction.\n"
                             "  Positive CMF = institutions quietly accumulating (bullish).\n"
                             "  Negative CMF = institutions distributing (bearish).\n"
                             "  ACTION: Prepare your entry order. Direction lean predicts breakout side.\n\n"
                             "Phase 3 — EXPLOSION\n"
                             "  Price closes above upper band (bullish) or below lower band (bearish).\n"
                             "  BBW starts expanding rapidly. This is the BUY/SHORT candle.\n"
                             "  ACTION: If all 5 conditions met → ENTER. Check for head fake first.\n\n"
                             "NORMAL = Not in squeeze (no setup)\n"
                             "POST-BREAKOUT = Trend is running after the breakout"),
    ("Phase Code",       12, "Numeric phase code for filtering:\n"
                             "  1 = COMPRESSION (spring coiling)\n"
                             "  2 = DIRECTION CLUES (positioning)\n"
                             "  3 = EXPLOSION (breakout)\n"
                             "  0 = NORMAL (no squeeze)\n"
                             "  4 = POST-BREAKOUT (trend running)"),
    ("Direction Lean",   14, "During a squeeze, which direction is the breakout likely to go?\n\n"
                             "BULLISH  = CMF positive + MFI > 60 + %b > 0.50\n"
                             "  → Upside breakout more likely. Prepare buy order above upper band.\n\n"
                             "BEARISH  = CMF negative + MFI < 40 + %b < 0.50\n"
                             "  → Downside breakout more likely. Stay away or prepare for short.\n\n"
                             "NEUTRAL  = Indicators balanced. No clear lean.\n"
                             "  → Wait for more clarity. Do not guess direction."),
    ("Squeeze Days",     13, "How many consecutive calendar days the stock has been in active squeeze.\n\n"
                             "Interpretation:\n"
                             "  1-5 days   = Fresh squeeze. Spring just started coiling.\n"
                             "  6-10 days  = Moderate energy stored. Watch closely.\n"
                             "  11-20 days = High energy. Breakout could be significant.\n"
                             "  20+ days   = Very high energy. Potentially large move when released.\n\n"
                             "The longer the squeeze, the BIGGER the expected breakout."),

    # ── Group A: Bollinger Bands ───────────────────────────────────
    ("Price (Close) ₹",  14, "Last closing price of the stock (NSE, 3:30 PM IST).\n"
                             "This is the price on which ALL indicators are calculated."),
    ("BB Upper ₹",       12, "Upper Bollinger Band value.\n"
                             "Formula: 20-day SMA + (2 × 20-day standard deviation)\n"
                             "Meaning: The statistical upper boundary of 'normal' price range.\n"
                             "Condition 2 requires price to CLOSE above this level."),
    ("BB Middle ₹",      12, "Middle Bollinger Band = 20-day Simple Moving Average of Close.\n"
                             "This is the trend baseline. Price above = bullish. Price below = bearish.\n"
                             "Also used as HOLD condition: price must stay above this."),
    ("BB Lower ₹",       12, "Lower Bollinger Band value.\n"
                             "Formula: 20-day SMA - (2 × 20-day standard deviation)\n"
                             "Exit Signal 2: If price touches or goes below this = exit the trade\n"
                             "(the full move from upper to lower band is the 'textbook profit target')."),
    ("SAR ₹",            11, "Parabolic SAR (Stop And Reverse) value in rupees.\n\n"
                             "Parameters used: Init AF=0.02, Step=0.02, Max=0.20\n\n"
                             "HOW IT WORKS:\n"
                             "  If SAR dots appear BELOW candles → uptrend → HOLD the trade.\n"
                             "  If SAR dots appear ABOVE candles → downtrend → EXIT immediately.\n\n"
                             "STOP LOSS: The SAR value shown IS your stop loss.\n"
                             "  If you are in a BUY trade, exit the MOMENT price closes below this value.\n"
                             "  The SAR automatically trails upward as the stock price rises."),
    ("SAR Trend",        12, "UPTREND  = SAR dots are below the candles. Stay in the trade.\n"
                             "DOWNTREND = SAR dots are above the candles. Do NOT buy / EXIT existing trade.\n\n"
                             "This is Exit Signal 1 (the PRIMARY exit). The moment SAR flips from\n"
                             "UPTREND to DOWNTREND = exit at next morning's open. No exceptions."),

    # ── Group A: Volume ────────────────────────────────────────────
    ("Volume",           14, "Total shares traded on the last trading day.\n"
                             "A raw number, e.g. 12,345,678 = 1.23 crore shares traded.\n"
                             "CRITICAL: Volume must be ABOVE the 50-day SMA for Condition 3."),
    ("Vol SMA50",        14, "50-day Simple Moving Average of daily volume.\n"
                             "This is the 'yellow reference line' on the volume panel.\n"
                             "It represents the stock's NORMAL trading volume.\n\n"
                             "If today's volume > Vol SMA50 → above average → institutional participation."),
    ("Vol Ratio",         11, "Today's Volume ÷ 50-day Volume SMA.\n\n"
                             "Interpretation:\n"
                             "  < 1.0 = Below average volume → weak signal → possible head fake\n"
                             "  1.0 - 1.5 = Average volume → decent confirmation\n"
                             "  1.5 - 3.0 = Strong volume → good confirmation → institutions participating\n"
                             "  > 3.0 = Very high volume → very strong breakout or news event"),

    # ── Group B: BBW / %b ─────────────────────────────────────────
    ("BBW",              10, "BandWidth = (Upper Band - Lower Band) ÷ Middle Band\n\n"
                             "What it measures: How TIGHT or WIDE the Bollinger Bands are.\n\n"
                             "Interpretation:\n"
                             "  Low BBW (e.g. 0.05 - 0.08) = Bands very tight = LOW volatility = SQUEEZE\n"
                             "  High BBW (e.g. 0.20 - 0.40) = Bands wide = HIGH volatility = post-breakout\n\n"
                             "SQUEEZE is SET when current BBW is at or near its 6-month rolling minimum.\n"
                             "This is CONDITION 1 — the foundation of the entire strategy."),
    ("BBW 6M Min",       11, "Rolling 6-month minimum of BandWidth (lookback = 126 trading days).\n"
                             "This is the DYNAMIC squeeze threshold — unique to each stock.\n\n"
                             "When current BBW ≤ (6M Min × 1.05) → Squeeze is ON.\n\n"
                             "Why dynamic? Because different stocks have different volatility profiles.\n"
                             "A BBW of 0.08 may be 'squeezed' for Reliance but 'wide' for a penny stock."),
    ("Squeeze ON",       11, "YES = The stock is currently in an active squeeze (Condition 1 met).\n"
                             "NO  = Not in squeeze. No setup.\n\n"
                             "Squeeze ON means: BBW ≤ BBW_6M_Min × 1.05  OR  BBW ≤ 0.08 (absolute)\n\n"
                             "This is required for ALL signal types (BUY, WAIT, HEAD FAKE)."),
    ("%b (Percent B)",   12, "Where is today's closing price within the Bollinger Bands?\n\n"
                             "Formula: %b = (Close - Lower Band) ÷ (Upper Band - Lower Band)\n\n"
                             "Values:\n"
                             "  0.00 = Price exactly at lower band\n"
                             "  0.50 = Price at the middle (20-day SMA)\n"
                             "  1.00 = Price at upper band\n"
                             "  > 1.00 = Price ABOVE upper band (BREAKOUT)\n"
                             "  < 0.00 = Price BELOW lower band (BREAKDOWN)\n\n"
                             "Used for Phase 2 direction lean analysis."),

    # ── Group C: CMF / MFI ────────────────────────────────────────
    ("CMF",              10, "Chaikin Money Flow — the 'smart money detector'.\n\n"
                             "Period: 20 days\n"
                             "Range: -1.0 to +1.0 (shown as decimal, e.g. +0.134)\n\n"
                             "Formula: Σ(MFV, 20) ÷ Σ(Volume, 20)\n"
                             "where MFV = ((Close-Low)-(High-Close))/(High-Low) × Volume\n\n"
                             "What it measures: Are institutional players buying (CMF +) or selling (CMF -)?.\n"
                             "The secret: CMF can be POSITIVE while price is flat during a squeeze.\n"
                             "This means institutions are QUIETLY ACCUMULATING before the breakout.\n\n"
                             "Levels:\n"
                             "  > +0.10 = Strong accumulation → big players buying aggressively\n"
                             "  0 to +0.10 = Mild buying → lean bullish\n"
                             "  -0.10 to 0 = Mild selling → lean bearish\n"
                             "  < -0.10 = Strong distribution → big players selling aggressively"),
    ("MFI",              10, "Money Flow Index — the 'fuel gauge'.\n\n"
                             "Period: 10 days (half of BB period — John Bollinger's specification)\n"
                             "Range: 0 to 100\n\n"
                             "Formula: 100 - (100 / (1 + Positive_MF_Sum / Negative_MF_Sum))\n"
                             "where Typical Price = (High + Low + Close) / 3\n\n"
                             "What it measures: Buying vs selling pressure with volume weight.\n"
                             "Like RSI but includes volume — more reliable for breakout confirmation.\n\n"
                             "Levels (strategy-specific — NOT the usual 70/30):\n"
                             "  > 80 = Maximum fuel → Enter FULL position\n"
                             "  50 to 80 = Sufficient fuel → Enter normal position\n"
                             "  < 50 on breakout = Weak fuel → Skip or use half position\n"
                             "  < 20 = Extremely oversold (not a buy condition here)"),

    # ── 5 Buy Conditions ─────────────────────────────────────────
    ("C1: Squeeze ON",   13, "Condition 1: Is the squeeze SET?\n"
                             "✅ YES = BBW is at or below 6-month rolling minimum. Spring is coiled.\n"
                             "❌ NO  = BBW too high. No setup. Skip this stock.\n"
                             "Weight in confidence score: 25 points."),
    ("C2: Price Break",  13, "Condition 2: Has price CLOSED above the upper Bollinger Band?\n"
                             "✅ YES = Closing price > Upper Band. Spring is being released upward.\n"
                             "❌ NO  = Price still inside the bands. No confirmed breakout yet.\n"
                             "IMPORTANT: Must be a CLOSE above the band, not just an intraday wick.\n"
                             "Weight in confidence score: 25 points."),
    ("C3: Volume OK",    13, "Condition 3: Is volume confirming the breakout?\n"
                             "✅ YES = Today's candle is green (close ≥ prev close) AND volume > 50-day SMA.\n"
                             "❌ NO  = Either candle is red or volume is below average.\n"
                             "If volume < average on a breakout = likely HEAD FAKE. Eliminate 80% of fakes.\n"
                             "Weight in confidence score: 20 points."),
    ("C4: CMF > 0",      13, "Condition 4: Is smart money flowing INTO the stock?\n"
                             "✅ YES = CMF > 0 (positive). Institutions are accumulating.\n"
                             "❌ NO  = CMF ≤ 0. Smart money not participating. Skip this breakout.\n"
                             "Ideal: CMF > +0.10 (strong accumulation) = bonus 5 points.\n"
                             "Weight in confidence score: 15 points."),
    ("C5: MFI Rising",   13, "Condition 5: Does the breakout have buying fuel?\n"
                             "✅ YES = MFI > 50 AND MFI is higher than yesterday's MFI (rising).\n"
                             "❌ NO  = MFI ≤ 50 or MFI is falling. Breakout running on empty.\n"
                             "MFI > 80 on breakout = maximum fuel = bonus 5 points.\n"
                             "Weight in confidence score: 15 points."),

    # ── Exit signals ──────────────────────────────────────────────
    ("Exit: SAR Flip",   13, "Exit Signal 1 (PRIMARY — most urgent):\n"
                             "YES = Parabolic SAR has FLIPPED. Price closed below the SAR dot.\n"
                             "  → EXIT at TOMORROW'S market open (9:15 AM NSE). No exceptions.\n"
                             "  → The uptrend is officially over. Do not wait for recovery.\n"
                             "NO = SAR is still bullish. No primary exit trigger."),
    ("Exit: LowBand",    13, "Exit Signal 2 (MAXIMUM PROFIT EXIT):\n"
                             "YES = Price has touched or crossed the LOWER Bollinger Band.\n"
                             "  → The full move from upper band to lower band is COMPLETE.\n"
                             "  → This is the textbook profit-taking exit.\n"
                             "  → Exit at next morning's open. You've captured the full squeeze move.\n"
                             "NO = Price has not reached the lower band yet."),
    ("Exit: DblNeg",     13, "Exit Signal 3 (EARLY WARNING — least urgent but valuable):\n"
                             "YES = CMF < 0 AND MFI < 50 simultaneously. 'Double negative'.\n"
                             "  → Smart money is exiting (CMF negative)\n"
                             "  → Buying fuel exhausted (MFI below 50)\n"
                             "  → Breakout is losing steam. Exit before SAR flips to protect profits.\n"
                             "NO = CMF/MFI not both negative."),
    ("Head Fake",        12, "Is this breakout suspicious / likely fake?\n\n"
                             "YES = 2 or more of these warning signals present:\n"
                             "  1. Volume below 50-day SMA (eliminates 80% of fakes)\n"
                             "  2. CMF negative on upside breakout\n"
                             "  3. MFI below 50 on upside breakout\n"
                             "  4. BBW not expanding (bands not widening after supposed breakout)\n"
                             "  5. Long upper wick (> 60% of candle range) — rejection candle\n\n"
                             "NO = Breakout appears genuine.\n\n"
                             "ACTION on HEAD FAKE: DO NOT ENTER. Wait 2-3 days. The REAL move\n"
                             "often comes in the OPPOSITE direction as trapped buyers capitulate."),
    ("Stop Loss ₹",      12, "Recommended stop loss level = current Parabolic SAR value.\n"
                             "If you are in a BUY trade:\n"
                             "  → Check this value every evening after market close.\n"
                             "  → If next day's price closes BELOW this → exit immediately.\n"
                             "  → The SAR automatically moves UP as the stock price rises.\n"
                             "  → Never LOWER your stop loss. Only ever trail it upward."),

    # ── Fundamentals ──────────────────────────────────────────────
    ("Fund. Score /100", 14, "Fundamental strength score from 0 to 100.\n"
                             "Calculated from: P/E ratio, ROE %, Debt/Equity, Revenue growth,\n"
                             "Current Ratio, Profit Margin, and Earnings growth.\n\n"
                             "Interpretation:\n"
                             "  65-100 = STRONG FUNDAMENTALS → high conviction → full position\n"
                             "  40-64  = MODERATE FUNDAMENTALS → average business → normal position\n"
                             "  0-39   = WEAK FUNDAMENTALS → business struggling → reduce/skip\n\n"
                             "NOTE: Fundamentals do NOT affect buy/sell signals.\n"
                             "They only affect HOW MUCH you put in, not WHEN you enter."),
    ("Fund. Verdict",    22, "One-line fundamental verdict text, e.g.:\n"
                             "'STRONG FUNDAMENTALS — Good ROE, Low Debt, Growing Revenue'\n"
                             "'MODERATE FUNDAMENTALS — Average metrics'\n"
                             "'WEAK FUNDAMENTALS — High debt, poor margins'"),
    ("P/E Ratio",        10, "Price-to-Earnings Ratio (TTM = Trailing Twelve Months).\n"
                             "Formula: Current Share Price ÷ Earnings Per Share\n\n"
                             "Meaning: How many rupees are you paying for each rupee of profit.\n"
                             "P/E = 20 means you pay ₹20 for every ₹1 of annual earnings.\n\n"
                             "Benchmarks:\n"
                             "  < 15 = Cheap (but check if earnings are declining)\n"
                             "  15-25 = Reasonable (most large-caps trade here)\n"
                             "  > 25 = Expensive (requires strong growth justification)\n"
                             "  > 50 = Very expensive (only justified for very high-growth companies)"),
    ("ROE %",            10, "Return on Equity — how efficiently management uses shareholders' money.\n"
                             "Formula: Net Profit ÷ Shareholders' Equity × 100\n\n"
                             "Benchmarks:\n"
                             "  > 20% = Excellent (rare, strong management quality)\n"
                             "  15-20% = Good (above average)\n"
                             "  10-15% = Average\n"
                             "  < 10%  = Below average (poor capital allocation)\n\n"
                             "Warren Buffett looks for ROE > 15% consistently for 5+ years."),
    ("Debt/Equity",      11, "Total Debt ÷ Shareholders' Equity.\n\n"
                             "Measures financial risk / leverage.\n\n"
                             "Benchmarks:\n"
                             "  < 0.5  = Conservative (very safe, low risk)\n"
                             "  0.5-1.0 = Moderate (acceptable)\n"
                             "  1.0-2.0 = High debt (risky in downturns)\n"
                             "  > 2.0  = Very high debt (potential default risk)\n\n"
                             "Avoid high D/E stocks in rising interest rate environments."),
    ("Revenue Gr %",     12, "Year-over-year revenue growth percentage.\n"
                             "Example: 15.3% means revenue grew 15.3% compared to same period last year.\n\n"
                             "Benchmarks:\n"
                             "  > 20% = High growth (premium valuation justified)\n"
                             "  10-20% = Good growth\n"
                             "  5-10%  = Moderate growth\n"
                             "  < 5%   = Slow growth (watch for value trap)\n"
                             "  Negative = Revenue declining (high risk)"),
    ("Div Yield %",      11, "Annual Dividend ÷ Current Price × 100.\n"
                             "Example: 2.5% means ₹2.50 dividend per year for every ₹100 invested.\n\n"
                             "Benchmarks:\n"
                             "  > 3% = Good income (like a 3% annual return PLUS capital gains)\n"
                             "  1-3% = Moderate income\n"
                             "  < 1% = Low income (growth stock likely)\n"
                             "  0%   = No dividend (all profits reinvested for growth)"),
]


# ══════════════════════════════════════════════════════════════════
#  PHASE → NUMERIC CODE
# ══════════════════════════════════════════════════════════════════

PHASE_CODE = {
    "COMPRESSION":      1,
    "DIRECTION":        2,
    "EXPLOSION":        3,
    "NORMAL":           0,
    "POST-BREAKOUT":    4,
    "INSUFFICIENT_DATA": -1,
    "ERROR":            -1,
    "UNKNOWN":          -1,
}

PHASE_LABEL = {
    "COMPRESSION":   "Phase 1 — COMPRESSION (Spring Coiling)",
    "DIRECTION":     "Phase 2 — DIRECTION CLUES (Positioning)",
    "EXPLOSION":     "Phase 3 — EXPLOSION (Breakout)",
    "NORMAL":        "No Active Squeeze",
    "POST-BREAKOUT": "Post-Breakout Trend Running",
    "INSUFFICIENT_DATA": "Insufficient Data",
    "ERROR":         "Analysis Error",
    "UNKNOWN":       "Unknown",
}


# ══════════════════════════════════════════════════════════════════
#  SIGNAL → ROW FILL & FONT COLOUR
# ══════════════════════════════════════════════════════════════════

def _row_style(sig: SignalResult):
    """Returns (fill_hex, font_color_hex, signal_text) for a given signal."""
    if sig.buy_signal:
        return ROW_BUY,       DARK_GREEN, "🚀 BUY"
    if sig.sell_signal:
        return ROW_SELL,      DARK_RED,   "🔴 SELL / EXIT"
    if sig.hold_signal:
        return ROW_HOLD,      DARK_BLUE,  "🟢 HOLD"
    if sig.head_fake:
        return ROW_HEAD_FAKE, ORANGE,     "⚠️ HEAD FAKE"
    if sig.wait_signal:
        return ROW_WAIT,      BLACK,      "⏳ WAIT"
    if sig.phase in ("INSUFFICIENT_DATA", "ERROR"):
        return ROW_ERROR,     BLACK,      "❓ NO DATA"
    return ROW_MONITOR,   BLACK,      "⚪ MONITOR"


# ══════════════════════════════════════════════════════════════════
#  CELL VALUE HELPERS
# ══════════════════════════════════════════════════════════════════

def _yn(val: bool) -> str:
    return "YES" if val else "NO"


def _fmt(val, decimals=2) -> str:
    if val is None:
        return "N/A"
    try:
        return f"{float(val):.{decimals}f}"
    except Exception:
        return str(val)


def _vol_ratio(vol: float, sma: float) -> str:
    if sma <= 0:
        return "N/A"
    r = vol / sma
    return f"{r:.2f}x"


# ══════════════════════════════════════════════════════════════════
#  SHEET 1 — RESULTS
# ══════════════════════════════════════════════════════════════════

def _write_results_sheet(ws: Worksheet,
                         results: list[tuple[SignalResult, Optional[FundamentalData]]],
                         mode: str,
                         scan_date: str):
    """Write the main results table with colour coding and cell comments."""

    # ── Title row ────────────────────────────────────────────────
    mode_titles = {
        "BUY":     "BUY Signals — All 5 Conditions Met",
        "SELL":    "SELL / EXIT Signals — Exit Conditions Triggered",
        "SQUEEZE": "SQUEEZE Stocks — Phase 1 & 2 (Spring Coiling)",
        "ALL":     "Full Market Scan — All Signal Types",
        "SINGLE":  "Single Stock Analysis",
    }
    title = f"BOLLINGER BAND SQUEEZE STRATEGY — {mode_titles.get(mode, mode)}"
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.fill  = _fill(HDR_DARK)
    title_cell.font  = _font(bold=True, color=WHITE, size=13)
    title_cell.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 22

    # ── Subtitle ─────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    sub_cell = ws["A2"]
    sub_cell.value = (f"Scan Date: {scan_date}  |  Universe: NSE India  |  "
                      f"Based on: Bollinger on Bollinger Bands — John Bollinger CFA CMT  |  "
                      f"Stocks in this report: {len(results)}")
    sub_cell.fill  = _fill("2E4057")
    sub_cell.font  = _font(color=WHITE, size=9, italic=True)
    sub_cell.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 16

    # ── Group header row ─────────────────────────────────────────
    # Map each column header to its group
    group_map = {
        "Ticker": ("IDENTITY", HDR_META),
        "Company Name": ("IDENTITY", HDR_META),
        "Sector": ("IDENTITY", HDR_META),
        "Scan Date": ("IDENTITY", HDR_META),
        "Last Data Date": ("IDENTITY", HDR_META),
        "SIGNAL": ("SIGNAL", "8B0000"),
        "Confidence /100": ("SIGNAL", "8B0000"),
        "Action Message": ("SIGNAL", "8B0000"),
        "Phase": ("PHASE ANALYSIS", "1A5276"),
        "Phase Code": ("PHASE ANALYSIS", "1A5276"),
        "Direction Lean": ("PHASE ANALYSIS", "1A5276"),
        "Squeeze Days": ("PHASE ANALYSIS", "1A5276"),
        "Price (Close) ₹": ("GROUP A — PRICE & TREND", HDR_GROUP_A),
        "BB Upper ₹": ("GROUP A — PRICE & TREND", HDR_GROUP_A),
        "BB Middle ₹": ("GROUP A — PRICE & TREND", HDR_GROUP_A),
        "BB Lower ₹": ("GROUP A — PRICE & TREND", HDR_GROUP_A),
        "SAR ₹": ("GROUP A — PRICE & TREND", HDR_GROUP_A),
        "SAR Trend": ("GROUP A — PRICE & TREND", HDR_GROUP_A),
        "Volume": ("GROUP A — VOLUME", HDR_GROUP_A),
        "Vol SMA50": ("GROUP A — VOLUME", HDR_GROUP_A),
        "Vol Ratio": ("GROUP A — VOLUME", HDR_GROUP_A),
        "BBW": ("GROUP B — SQUEEZE", HDR_GROUP_B),
        "BBW 6M Min": ("GROUP B — SQUEEZE", HDR_GROUP_B),
        "Squeeze ON": ("GROUP B — SQUEEZE", HDR_GROUP_B),
        "%b (Percent B)": ("GROUP B — SQUEEZE", HDR_GROUP_B),
        "CMF": ("GROUP C — DIRECTION", HDR_GROUP_C),
        "MFI": ("GROUP C — DIRECTION", HDR_GROUP_C),
        "C1: Squeeze ON": ("5 BUY CONDITIONS", "7B241C"),
        "C2: Price Break": ("5 BUY CONDITIONS", "7B241C"),
        "C3: Volume OK": ("5 BUY CONDITIONS", "7B241C"),
        "C4: CMF > 0": ("5 BUY CONDITIONS", "7B241C"),
        "C5: MFI Rising": ("5 BUY CONDITIONS", "7B241C"),
        "Exit: SAR Flip": ("3 EXIT SIGNALS", "641E16"),
        "Exit: LowBand": ("3 EXIT SIGNALS", "641E16"),
        "Exit: DblNeg": ("3 EXIT SIGNALS", "641E16"),
        "Head Fake": ("3 EXIT SIGNALS", "641E16"),
        "Stop Loss ₹": ("3 EXIT SIGNALS", "641E16"),
        "Fund. Score /100": ("FUNDAMENTALS", HDR_FUND),
        "Fund. Verdict": ("FUNDAMENTALS", HDR_FUND),
        "P/E Ratio": ("FUNDAMENTALS", HDR_FUND),
        "ROE %": ("FUNDAMENTALS", HDR_FUND),
        "Debt/Equity": ("FUNDAMENTALS", HDR_FUND),
        "Revenue Gr %": ("FUNDAMENTALS", HDR_FUND),
        "Div Yield %": ("FUNDAMENTALS", HDR_FUND),
    }

    # Merge group header cells
    col_groups = {}
    for i, (hdr, _, _) in enumerate(COLUMNS, 1):
        grp, grp_color = group_map.get(hdr, ("OTHER", HDR_META))
        if grp not in col_groups:
            col_groups[grp] = {"start": i, "end": i, "color": grp_color}
        else:
            col_groups[grp]["end"] = i

    for grp, info in col_groups.items():
        start_col = get_column_letter(info["start"])
        end_col   = get_column_letter(info["end"])
        if info["start"] == info["end"]:
            cell = ws[f"{start_col}3"]
        else:
            ws.merge_cells(f"{start_col}3:{end_col}3")
            cell = ws[f"{start_col}3"]
        cell.value     = grp
        cell.fill      = _fill(info["color"])
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.alignment = _align("center", "center")
    ws.row_dimensions[3].height = 14

    # ── Column header row (row 4) ─────────────────────────────────
    for col_idx, (hdr, width, tooltip) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx, value=hdr)
        _, grp_color = group_map.get(hdr, ("OTHER", "2C3E50"))
        cell.fill      = _fill(grp_color)
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border    = THIN_BORDER
        _set_col_width(ws, col_idx, width)
        # Add comment/tooltip
        from openpyxl.comments import Comment
        comment = Comment(tooltip, "BB Squeeze Analyser")
        comment.width  = 320
        comment.height = 180
        cell.comment   = comment
    ws.row_dimensions[4].height = 30

    # Freeze panes at row 5 col B
    ws.freeze_panes = "C5"

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, (sig, fd) in enumerate(results, 5):
        fill_hex, font_hex, signal_text = _row_style(sig)
        row_fill = _fill(fill_hex)
        row_font = _font(color=font_hex, size=9)

        def cell_val(col_num, value):
            c = ws.cell(row=row_idx, column=col_num, value=value)
            c.fill      = row_fill
            c.font      = row_font
            c.alignment = _align("left", "center", wrap=False)
            c.border    = THIN_BORDER
            return c

        col = 1

        # Identity
        cell_val(col, sig.ticker);                                    col += 1
        cell_val(col, fd.company_name if fd and not fd.fetch_error else "N/A"); col += 1
        cell_val(col, fd.sector if fd and not fd.fetch_error else "N/A");       col += 1
        cell_val(col, scan_date);                                     col += 1
        cell_val(col, "");                                            col += 1   # last data date — filled below

        # Signal
        sig_cell = cell_val(col, signal_text)
        # Override signal cell colour for stronger visibility
        if sig.buy_signal:
            sig_cell.fill = _fill("00B050")
            sig_cell.font = _font(bold=True, color=WHITE, size=10)
        elif sig.sell_signal:
            sig_cell.fill = _fill("FF0000")
            sig_cell.font = _font(bold=True, color=WHITE, size=10)
        elif sig.hold_signal:
            sig_cell.fill = _fill("4472C4")
            sig_cell.font = _font(bold=True, color=WHITE, size=10)
        elif sig.wait_signal:
            sig_cell.fill = _fill("FFC000")
            sig_cell.font = _font(bold=True, color=BLACK, size=10)
        elif sig.head_fake:
            sig_cell.fill = _fill("E26B0A")
            sig_cell.font = _font(bold=True, color=WHITE, size=10)
        col += 1

        # Confidence — colour gradient green/yellow/red
        conf_cell = cell_val(col, sig.confidence)
        conf_cell.alignment = _align("center", "center")
        if sig.confidence >= 80:
            conf_cell.fill = _fill("00B050"); conf_cell.font = _font(bold=True, color=WHITE, size=9)
        elif sig.confidence >= 60:
            conf_cell.fill = _fill("FFC000"); conf_cell.font = _font(bold=True, color=BLACK, size=9)
        else:
            conf_cell.fill = _fill("FF0000"); conf_cell.font = _font(bold=True, color=WHITE, size=9)
        col += 1

        # Action message — word wrap
        msg_cell = ws.cell(row=row_idx, column=col,
                           value=sig.action_message.replace("\n", " "))
        msg_cell.fill      = row_fill
        msg_cell.font      = _font(color=font_hex, size=8, italic=True)
        msg_cell.alignment = _align("left", "center", wrap=True)
        msg_cell.border    = THIN_BORDER
        col += 1

        # Phase (full label)
        ph_cell = cell_val(col, PHASE_LABEL.get(sig.phase, sig.phase))
        # Phase colour
        phase_colors = {
            "COMPRESSION":   ("2471A3", WHITE),
            "DIRECTION":     ("D4AC0D", BLACK),
            "EXPLOSION":     ("CB4335", WHITE),
            "NORMAL":        ("AAB7B8", BLACK),
            "POST-BREAKOUT": ("1E8449", WHITE),
        }
        pc = phase_colors.get(sig.phase)
        if pc:
            ph_cell.fill = _fill(pc[0])
            ph_cell.font = _font(bold=True, color=pc[1], size=9)
        col += 1

        # Phase code
        pc_cell = cell_val(col, PHASE_CODE.get(sig.phase, -1))
        pc_cell.alignment = _align("center", "center")
        col += 1

        # Direction lean
        lean_cell = cell_val(col, sig.direction_lean)
        lean_cell.alignment = _align("center", "center")
        if sig.direction_lean == "BULLISH":
            lean_cell.fill = _fill("C6EFCE"); lean_cell.font = _font(bold=True, color=DARK_GREEN)
        elif sig.direction_lean == "BEARISH":
            lean_cell.fill = _fill("FFC7CE"); lean_cell.font = _font(bold=True, color=DARK_RED)
        col += 1

        # Squeeze days
        sqd_cell = cell_val(col, sig.squeeze_days if sig.squeeze_days > 0 else "")
        sqd_cell.alignment = _align("center", "center")
        if sig.squeeze_days >= 20:
            sqd_cell.fill = _fill("00B050"); sqd_cell.font = _font(bold=True, color=WHITE)
        elif sig.squeeze_days >= 10:
            sqd_cell.fill = _fill("FFC000"); sqd_cell.font = _font(bold=True, color=BLACK)
        col += 1

        # Price / BB / SAR
        for val in [sig.current_price, sig.bb_upper, sig.bb_mid, sig.bb_lower]:
            c = cell_val(col, round(val, 2));  c.alignment = _align("right", "center"); col += 1

        # SAR ₹
        sar_c = cell_val(col, round(sig.sar, 2)); sar_c.alignment = _align("right", "center"); col += 1

        # SAR Trend
        sar_trend = "UPTREND" if sig.sar_bull else "DOWNTREND"
        st_cell = cell_val(col, sar_trend)
        st_cell.alignment = _align("center", "center")
        if sig.sar_bull:
            st_cell.fill = _fill("C6EFCE"); st_cell.font = _font(bold=True, color=DARK_GREEN)
        else:
            st_cell.fill = _fill("FFC7CE"); st_cell.font = _font(bold=True, color=DARK_RED)
        col += 1

        # Volume
        vol_c = cell_val(col, int(sig.volume)); vol_c.alignment = _align("right", "center"); col += 1

        # Vol SMA50
        vs_c = cell_val(col, int(sig.vol_sma50)); vs_c.alignment = _align("right", "center"); col += 1

        # Vol Ratio
        vr_val = (sig.volume / sig.vol_sma50) if sig.vol_sma50 > 0 else 0
        vr_cell = cell_val(col, round(vr_val, 2))
        vr_cell.alignment = _align("center", "center")
        if vr_val >= 1.5:
            vr_cell.fill = _fill("C6EFCE"); vr_cell.font = _font(bold=True, color=DARK_GREEN)
        elif vr_val >= 1.0:
            vr_cell.fill = _fill("EBF5FB")
        else:
            vr_cell.fill = _fill("FFC7CE"); vr_cell.font = _font(color=DARK_RED)
        col += 1

        # BBW
        bbw_cell = cell_val(col, round(sig.bbw, 4))
        bbw_cell.alignment = _align("center", "center")
        if sig.cond1_squeeze_on:
            bbw_cell.fill = _fill("C6EFCE"); bbw_cell.font = _font(bold=True, color=DARK_GREEN)
        col += 1

        # BBW 6M Min
        cell_val(col, round(sig.bbw_6m_min, 4)).alignment = _align("center", "center"); col += 1

        # Squeeze ON
        sq_cell = cell_val(col, _yn(sig.cond1_squeeze_on))
        sq_cell.alignment = _align("center", "center")
        sq_cell.fill = _fill("C6EFCE") if sig.cond1_squeeze_on else _fill("FFC7CE")
        sq_cell.font = _font(bold=True,
                             color=DARK_GREEN if sig.cond1_squeeze_on else DARK_RED)
        col += 1

        # %b
        pb_cell = cell_val(col, round(sig.percent_b, 3))
        pb_cell.alignment = _align("center", "center")
        if sig.percent_b > 1.0:
            pb_cell.fill = _fill("00B050"); pb_cell.font = _font(bold=True, color=WHITE)
        elif sig.percent_b > 0.5:
            pb_cell.fill = _fill("C6EFCE")
        elif sig.percent_b < 0.2:
            pb_cell.fill = _fill("FFC7CE")
        col += 1

        # CMF
        cmf_cell = cell_val(col, round(sig.cmf, 4))
        cmf_cell.alignment = _align("center", "center")
        if sig.cmf > 0.10:
            cmf_cell.fill = _fill("00B050"); cmf_cell.font = _font(bold=True, color=WHITE)
        elif sig.cmf > 0:
            cmf_cell.fill = _fill("C6EFCE")
        elif sig.cmf < -0.10:
            cmf_cell.fill = _fill("FF0000"); cmf_cell.font = _font(bold=True, color=WHITE)
        else:
            cmf_cell.fill = _fill("FFC7CE")
        col += 1

        # MFI
        mfi_cell = cell_val(col, round(sig.mfi, 1))
        mfi_cell.alignment = _align("center", "center")
        if sig.mfi > 80:
            mfi_cell.fill = _fill("00B050"); mfi_cell.font = _font(bold=True, color=WHITE)
        elif sig.mfi > 50:
            mfi_cell.fill = _fill("C6EFCE")
        elif sig.mfi < 20:
            mfi_cell.fill = _fill("FF0000"); mfi_cell.font = _font(bold=True, color=WHITE)
        else:
            mfi_cell.fill = _fill("FFC7CE")
        col += 1

        # 5 Conditions (YES/NO with green/red)
        for val in [sig.cond1_squeeze_on, sig.cond2_price_above,
                    sig.cond3_volume_ok, sig.cond4_cmf_positive, sig.cond5_mfi_above_50]:
            c = cell_val(col, _yn(val))
            c.alignment = _align("center", "center")
            c.fill = _fill("C6EFCE") if val else _fill("FFC7CE")
            c.font = _font(bold=True, color=DARK_GREEN if val else DARK_RED)
            col += 1

        # 3 Exits
        for val in [sig.exit_sar_flip, sig.exit_lower_band_tag, sig.exit_double_neg]:
            c = cell_val(col, _yn(val))
            c.alignment = _align("center", "center")
            c.fill = _fill("FFC7CE") if val else _fill("C6EFCE")
            c.font = _font(bold=val, color=DARK_RED if val else DARK_GREEN)
            col += 1

        # Head Fake
        hf_cell = cell_val(col, _yn(sig.head_fake))
        hf_cell.alignment = _align("center", "center")
        hf_cell.fill = _fill("FF0000") if sig.head_fake else _fill("C6EFCE")
        hf_cell.font = _font(bold=sig.head_fake,
                             color=WHITE if sig.head_fake else DARK_GREEN)
        col += 1

        # Stop Loss
        sl_cell = cell_val(col, round(sig.stop_loss, 2))
        sl_cell.alignment = _align("right", "center")
        if sig.stop_loss > 0:
            sl_cell.fill = _fill("FFEB9C")
            sl_cell.font = _font(bold=True, color="7B241C")
        col += 1

        # Fundamentals
        if fd and not fd.fetch_error:
            # Score
            fs_cell = cell_val(col, fd.fundamental_score)
            fs_cell.alignment = _align("center", "center")
            if fd.fundamental_score >= 65:
                fs_cell.fill = _fill("00B050"); fs_cell.font = _font(bold=True, color=WHITE)
            elif fd.fundamental_score >= 40:
                fs_cell.fill = _fill("FFC000"); fs_cell.font = _font(bold=True, color=BLACK)
            else:
                fs_cell.fill = _fill("FF0000"); fs_cell.font = _font(bold=True, color=WHITE)
            col += 1

            cell_val(col, fd.fundamental_verdict); col += 1
            cell_val(col, _fmt(fd.pe_ratio)).alignment  = _align("center", "center"); col += 1
            cell_val(col, _fmt(fd.roe)).alignment       = _align("center", "center"); col += 1
            cell_val(col, _fmt(fd.debt_to_equity)).alignment = _align("center", "center"); col += 1
            cell_val(col, _fmt(fd.revenue_growth)).alignment = _align("center", "center"); col += 1
            cell_val(col, _fmt(fd.dividend_yield)).alignment = _align("center", "center"); col += 1
        else:
            for _ in range(7):
                cell_val(col, "N/A"); col += 1

        ws.row_dimensions[row_idx].height = 28

    # ── Auto-filter ───────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A4:{get_column_letter(len(COLUMNS))}{4 + len(results)}"
    )


# ══════════════════════════════════════════════════════════════════
#  SHEET 2 — LEGEND
# ══════════════════════════════════════════════════════════════════

def _write_legend_sheet(ws: Worksheet):
    """Full glossary of every column, colour, and indicator."""
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 80
    ws.row_dimensions[1].height = 26

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "LEGEND — FULL COLUMN GLOSSARY"
    t.fill      = _fill(LEG_HEADER)
    t.font      = _font(bold=True, color=WHITE, size=13)
    t.alignment = _align("center", "center")

    # ── PART 1: Row Colour Legend ──────────────────────────────────
    ws.merge_cells("A2:C2")
    ws["A2"].value     = "SECTION 1 — ROW COLOUR CODES (what each row background colour means)"
    ws["A2"].fill      = _fill(LEG_SUBHEAD)
    ws["A2"].font      = _font(bold=True, color=WHITE, size=10)
    ws["A2"].alignment = _align("center", "center")

    color_rows = [
        ("SIGNAL",        "Row Background",  "What It Means",           "What You Do"),
        ("🚀 BUY",        ROW_BUY,          "Pale Green",               "All 5 buy conditions met. Enter tomorrow at 9:15 AM NSE open."),
        ("🔴 SELL/EXIT",  ROW_SELL,         "Pale Red",                 "Exit condition triggered. Exit at tomorrow's market open."),
        ("🟢 HOLD",       ROW_HOLD,         "Pale Blue",                "Trend intact. Stay in the trade. Trail stop loss to SAR."),
        ("⏳ WAIT",       ROW_WAIT,         "Pale Yellow/Amber",        "Squeeze set but no breakout. Watch daily for the explosion."),
        ("⚠️ HEAD FAKE",  ROW_HEAD_FAKE,    "Orange/Amber",             "DO NOT enter. False breakout. Wait for real move."),
        ("⚪ MONITOR",    ROW_MONITOR,      "Light Grey",               "No setup. Keep on watchlist."),
        ("❓ NO DATA",    ROW_ERROR,        "Grey",                     "Insufficient data to analyse. Need more history."),
    ]

    row = 3
    for i, (sig, hex_c, color_name, meaning) in enumerate(color_rows):
        if i == 0:  # header
            for ci, v in enumerate([sig, color_name, meaning], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = _fill("2C3E50"); c.font = _font(bold=True, color=WHITE, size=9)
                c.alignment = _align("center", "center"); c.border = THIN_BORDER
        else:
            ws.cell(row=row, column=1, value=sig).border = THIN_BORDER
            ws.cell(row=row, column=1).font = _font(bold=True, size=9)
            color_swatch = ws.cell(row=row, column=2, value=f"  {color_name}")
            color_swatch.fill = _fill(hex_c); color_swatch.border = THIN_BORDER
            ws.cell(row=row, column=3, value=meaning).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = _align("left", "center", wrap=True)
            ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # ── PART 2: Cell Colour Codes ─────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value     = "SECTION 2 — INDIVIDUAL CELL COLOUR CODES"
    ws[f"A{row}"].fill      = _fill(LEG_SUBHEAD)
    ws[f"A{row}"].font      = _font(bold=True, color=WHITE, size=10)
    ws[f"A{row}"].alignment = _align("center", "center")
    row += 1

    cell_colors = [
        ("Column",              "Cell Color",    "Meaning"),
        ("Signal",              "Bright Green",  "BUY signal — all 5 conditions met"),
        ("Signal",              "Bright Red",    "SELL/EXIT signal"),
        ("Signal",              "Blue",          "HOLD signal"),
        ("Signal",              "Orange",        "HEAD FAKE — do not enter"),
        ("Confidence",          "Green",         ">= 80/100 — high confidence — full position"),
        ("Confidence",          "Yellow",        "60-79/100 — moderate — half position"),
        ("Confidence",          "Red",           "< 60/100 — low confidence"),
        ("SAR Trend",           "Green",         "UPTREND — SAR dots below candles — stay in"),
        ("SAR Trend",           "Red",           "DOWNTREND — SAR dots above candles — exit"),
        ("Vol Ratio",           "Green",         ">= 1.5x — volume strongly above average"),
        ("Vol Ratio",           "Red",           "< 1.0x — below-average volume — weak"),
        ("BBW",                 "Green",         "Squeeze is ON (BBW at/below trigger)"),
        ("Squeeze ON",          "Green",         "YES — spring is coiled"),
        ("Squeeze ON",          "Red",           "NO — not in squeeze"),
        ("%b",                  "Bright Green",  "> 1.0 — price above upper band (breakout)"),
        ("%b",                  "Pale Green",    "0.5 to 1.0 — lean bullish zone"),
        ("%b",                  "Pale Red",      "< 0.2 — near bottom of band (bearish)"),
        ("CMF",                 "Bright Green",  "> +0.10 — strong accumulation"),
        ("CMF",                 "Pale Green",    "0 to +0.10 — mild buying"),
        ("CMF",                 "Pale Red",      "-0.10 to 0 — mild selling"),
        ("CMF",                 "Bright Red",    "< -0.10 — strong distribution"),
        ("MFI",                 "Bright Green",  "> 80 — maximum fuel"),
        ("MFI",                 "Pale Green",    "50 to 80 — sufficient fuel"),
        ("MFI",                 "Pale Red",      "< 50 — weak fuel"),
        ("MFI",                 "Bright Red",    "< 20 — extremely oversold"),
        ("C1 to C5",            "Green",         "Condition MET"),
        ("C1 to C5",            "Red",           "Condition NOT met"),
        ("Exit signals",        "Red",           "Exit triggered — action required"),
        ("Exit signals",        "Green",         "No exit trigger"),
        ("Head Fake",           "Red",           "Head fake detected — DO NOT enter"),
        ("Squeeze Days",        "Green",         ">= 20 days — very high energy stored"),
        ("Squeeze Days",        "Yellow",        "10-19 days — high energy"),
        ("Fund. Score",         "Green",         ">= 65 — strong fundamentals"),
        ("Fund. Score",         "Yellow",        "40-64 — moderate fundamentals"),
        ("Fund. Score",         "Red",           "< 40 — weak fundamentals"),
        ("Phase",               "Blue",          "Phase 1: COMPRESSION (coiling)"),
        ("Phase",               "Yellow",        "Phase 2: DIRECTION CLUES (positioning)"),
        ("Phase",               "Red",           "Phase 3: EXPLOSION (breakout)"),
        ("Phase",               "Green",         "POST-BREAKOUT (trend running)"),
    ]

    for i, (col_name, cell_col, meaning) in enumerate(cell_colors):
        alt_fill = LEG_ROW_ALT if i % 2 == 0 else WHITE
        ws.cell(row=row, column=1, value=col_name).fill    = _fill(alt_fill)
        ws.cell(row=row, column=1).font   = _font(size=9, bold=(i == 0))
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=cell_col).fill    = _fill(alt_fill)
        ws.cell(row=row, column=2).font   = _font(size=9, bold=(i == 0))
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=meaning).fill     = _fill(alt_fill)
        ws.cell(row=row, column=3).font   = _font(size=9, bold=(i == 0))
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = _align("left", "center", wrap=True)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1

    # ── PART 3: Column-by-column glossary ─────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value     = "SECTION 3 — FULL COLUMN GLOSSARY (indicator formulas, levels, and meaning)"
    ws[f"A{row}"].fill      = _fill(LEG_SUBHEAD)
    ws[f"A{row}"].font      = _font(bold=True, color=WHITE, size=10)
    ws[f"A{row}"].alignment = _align("center", "center")
    row += 1

    for i, (hdr, _, tooltip) in enumerate(COLUMNS):
        alt_fill = LEG_ROW_ALT if i % 2 == 0 else WHITE
        ws.cell(row=row, column=1, value=hdr).fill    = _fill(alt_fill)
        ws.cell(row=row, column=1).font   = _font(size=9, bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = _align("left", "top")
        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row=row, column=2, value=tooltip).fill = _fill(alt_fill)
        ws.cell(row=row, column=2).font   = _font(size=8)
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = _align("left", "top", wrap=True)
        lines = tooltip.count("\n") + 1
        ws.row_dimensions[row].height = max(14, min(lines * 12, 80))
        row += 1


# ══════════════════════════════════════════════════════════════════
#  SHEET 3 — PHASE GUIDE
# ══════════════════════════════════════════════════════════════════

def _write_phase_sheet(ws: Worksheet):
    """Detailed explanation of all 3 squeeze phases."""
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "SQUEEZE PHASE ANALYSIS GUIDE — The 3-Phase Bollinger Squeeze Model"
    t.fill      = _fill(HDR_DARK)
    t.font      = _font(bold=True, color=WHITE, size=13)
    t.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 24

    phases = [
        (
            "PHASE 1 — COMPRESSION",
            "2471A3",
            [
                ("Also Called",         "The 'Spring Coiling' phase"),
                ("Phase Code",          "1"),
                ("What's Happening",    "The Bollinger Bands are at their narrowest point in 6 months.\n"
                                        "Volatility has completely collapsed. The stock is moving sideways.\n"
                                        "Most traders are IGNORING this stock — there's nothing exciting happening."),
                ("BBW Status",          "BBW is at or near its 6-month rolling minimum (lowest value).\n"
                                        "This IS the squeeze. The bands are maximally compressed."),
                ("CMF Status",          "CMF is near zero — no clear money flow direction yet.\n"
                                        "Institutions may be quietly building positions, but not obvious yet."),
                ("MFI Status",          "MFI is near 50 — perfectly balanced buying and selling pressure."),
                ("%b Position",         "%b is near 0.50 — price sitting right at the 20-day moving average."),
                ("What To Do",          "1. Add this stock to your WATCHLIST immediately.\n"
                                        "2. Set a price alert just above the Upper Bollinger Band.\n"
                                        "3. Note how many days it's been in squeeze (Squeeze Days column).\n"
                                        "4. The LONGER the squeeze, the BIGGER the eventual breakout.\n"
                                        "5. Do NOT enter yet — this could continue for days or weeks."),
                ("Key Insight",         "John Bollinger: 'When the bands are as tight as they can get,\n"
                                        "get ready. The stock is about to make a significant move.\n"
                                        "The spring has been coiled. It WILL be released.'"),
            ]
        ),
        (
            "PHASE 2 — DIRECTION CLUES",
            "D4AC0D",
            [
                ("Also Called",         "The 'Positioning' or 'Pre-Breakout' phase"),
                ("Phase Code",          "2"),
                ("What's Happening",    "The stock is STILL in squeeze (bands still tight).\n"
                                        "BUT the direction indicators (CMF and MFI) are starting to move.\n"
                                        "Institutions are positioning themselves before the public sees anything."),
                ("CMF Status",          "CMF is moving away from zero:\n"
                                        "  Positive CMF (+) = Institutions BUYING quietly = BULLISH lean\n"
                                        "  Negative CMF (-) = Institutions SELLING quietly = BEARISH lean\n"
                                        "This is the KEY signal that reveals what big money is doing."),
                ("MFI Status",          "MFI is moving away from 50:\n"
                                        "  MFI > 60 = More buying pressure building = BULLISH\n"
                                        "  MFI < 40 = More selling pressure building = BEARISH"),
                ("%b Position",         "%b is drifting from 0.50:\n"
                                        "  %b approaching 0.80-1.0 = price moving toward upper band = BULLISH\n"
                                        "  %b approaching 0.20-0.0 = price moving toward lower band = BEARISH"),
                ("Direction Lean",      "The software calculates a DIRECTION LEAN score:\n"
                                        "  BULLISH = upside breakout more likely (prepare BUY order)\n"
                                        "  BEARISH = downside breakdown more likely (stay away)\n"
                                        "  NEUTRAL = no clear direction yet (continue waiting)"),
                ("What To Do",          "1. Check the Direction Lean column.\n"
                                        "2. If BULLISH lean: Prepare your BUY order above the upper band.\n"
                                        "3. Decide your position size based on Confidence Score.\n"
                                        "4. Know your stop loss in advance (= current SAR value).\n"
                                        "5. Still do NOT enter — wait for the actual breakout (Phase 3)."),
                ("Key Insight",         "This phase can last just 1-2 days or up to 1-2 weeks.\n"
                                        "The more STRONGLY CMF is positive during Phase 2,\n"
                                        "the MORE POWERFUL the eventual upside breakout will be."),
            ]
        ),
        (
            "PHASE 3 — EXPLOSION (BREAKOUT)",
            "CB4335",
            [
                ("Also Called",         "The 'Entry' phase — this is the TRADE"),
                ("Phase Code",          "3"),
                ("What's Happening",    "Price has CLOSED above the upper Bollinger Band (bullish).\n"
                                        "OR Price has CLOSED below the lower Bollinger Band (bearish).\n"
                                        "The BBW is starting to EXPAND — bands are widening.\n"
                                        "This is the spring being RELEASED."),
                ("Entry Condition",     "You enter ONLY when ALL 5 conditions are simultaneously met:\n"
                                        "  ✅ C1: BBW at squeeze trigger\n"
                                        "  ✅ C2: Price closed ABOVE upper band\n"
                                        "  ✅ C3: Volume green AND above 50-day SMA\n"
                                        "  ✅ C4: CMF positive (smart money flowing in)\n"
                                        "  ✅ C5: MFI above 50 and RISING"),
                ("Head Fake Risk",      "NOT every price break above the upper band is real.\n"
                                        "A HEAD FAKE occurs when price spikes above but:\n"
                                        "  - Volume is BELOW average (no institutional participation)\n"
                                        "  - CMF is negative (smart money actually selling)\n"
                                        "  - MFI is below 50 (no buying fuel)\n"
                                        "The software detects head fakes. If HEAD FAKE = YES: do NOT enter."),
                ("Entry Timing",        "Signal is based on the day's CLOSING price.\n"
                                        "Enter at NEXT DAY'S market open (9:15 AM NSE).\n"
                                        "Do NOT enter intraday — wait for the close confirmation."),
                ("Stop Loss",           "Immediately set stop loss = Parabolic SAR value shown.\n"
                                        "Example: If SAR = ₹2,450 and you buy at ₹2,580:\n"
                                        "  Max risk = ₹2,580 - ₹2,450 = ₹130 per share (5% risk).\n"
                                        "  If price closes below ₹2,450 on ANY day → EXIT NEXT MORNING."),
                ("What To Do",          "1. If SIGNAL = BUY → Enter at tomorrow's 9:15 AM open.\n"
                                        "2. Set stop loss = SAR value (column Stop Loss ₹).\n"
                                        "3. Note your entry price and stop loss in a trade journal.\n"
                                        "4. Check the stock every evening after market close.\n"
                                        "5. The moment any exit signal triggers → exit next morning."),
                ("Key Insight",         "The typical move in a successful breakout is from the\n"
                                        "upper band ALL THE WAY to the lower band (or more).\n"
                                        "Target = Lower Bollinger Band. Stop = Parabolic SAR."),
            ]
        ),
        (
            "POST-BREAKOUT — TREND RUNNING",
            "1E8449",
            [
                ("Phase Code",          "4"),
                ("What's Happening",    "The trade is in progress. Price is trending after the breakout.\n"
                                        "SAR dots are below the candles — uptrend is intact.\n"
                                        "BBW is expanding as volatility returns to the stock."),
                ("Hold Conditions",     "HOLD when ALL of these are true:\n"
                                        "  ✅ SAR dots below candles (uptrend)\n"
                                        "  ✅ Price above the 20-day moving average (middle band)\n"
                                        "  ✅ CMF positive (smart money still holding)\n"
                                        "  ✅ MFI > 40 (fuel still present)"),
                ("Exit Conditions",     "EXIT when ANY ONE of these triggers:\n"
                                        "  🔴 Exit 1: Price closes below SAR dot (primary)\n"
                                        "  🔴 Exit 2: Price touches/crosses lower BB (max profit)\n"
                                        "  🟡 Exit 3: CMF<0 AND MFI<50 simultaneously (early warning)"),
                ("Trailing Stop",       "Every evening: look up the current SAR value.\n"
                                        "That is your new stop loss. The SAR automatically rises\n"
                                        "as price rises — locking in more profit each day.\n"
                                        "NEVER move your stop loss DOWN. Only trail it UP."),
                ("What To Do",          "Check the stock every evening after NSE closes (3:30 PM).\n"
                                        "If SIGNAL column shows HOLD → no action needed.\n"
                                        "If SIGNAL column shows SELL → exit next morning (9:15 AM)."),
            ]
        ),
    ]

    row = 2
    for phase_title, phase_color, rows in phases:
        # Phase header
        ws.merge_cells(f"A{row}:C{row}")
        ph = ws[f"A{row}"]
        ph.value     = phase_title
        ph.fill      = _fill(phase_color)
        ph.font      = _font(bold=True, color=WHITE, size=12)
        ph.alignment = _align("center", "center")
        ws.row_dimensions[row].height = 22
        row += 1

        for field_name, explanation in rows:
            ws.cell(row=row, column=1, value=field_name).font = _font(bold=True, size=9)
            ws.cell(row=row, column=1).fill  = _fill("EAF4FC")
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = _align("left", "top")
            ws.merge_cells(f"B{row}:C{row}")
            exp_cell = ws.cell(row=row, column=2, value=explanation)
            exp_cell.font      = _font(size=9)
            exp_cell.fill      = _fill("FDFEFE")
            exp_cell.border    = THIN_BORDER
            exp_cell.alignment = _align("left", "top", wrap=True)
            lines = explanation.count("\n") + 1
            ws.row_dimensions[row].height = max(14, min(lines * 13, 90))
            row += 1

        row += 1  # blank row between phases


# ══════════════════════════════════════════════════════════════════
#  SHEET 4 — HOW TO READ
# ══════════════════════════════════════════════════════════════════

def _write_how_to_sheet(ws: Worksheet):
    """Step-by-step reading guide."""
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "HOW TO READ THIS REPORT — Step-by-Step Guide for Every Menu Option"
    t.fill      = _fill(HDR_DARK)
    t.font      = _font(bold=True, color=WHITE, size=13)
    t.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 24

    sections = [
        ("STEP-BY-STEP: HOW TO READ THE BUY SIGNALS SHEET", "8B0000", [
            ("Step 1", "Look at the SIGNAL column first",
             "Find all rows highlighted in GREEN with '🚀 BUY'. These are the actionable stocks for tomorrow."),
            ("Step 2", "Check Confidence /100",
             "Sort by Confidence (highest to lowest). Higher confidence = stronger signal.\n"
             "85-100: Enter full position. 70-84: Enter 75%. 60-69: Enter 50%."),
            ("Step 3", "Read the Action Message",
             "The Action Message column tells you exactly what to do, your entry point (tomorrow's open),\n"
             "and your stop loss level (₹ value of Parabolic SAR)."),
            ("Step 4", "Check Stop Loss ₹",
             "This is the Parabolic SAR value = your stop loss.\n"
             "If tomorrow's closing price goes below this → exit the next morning."),
            ("Step 5", "Check Head Fake column",
             "Even for BUY signals, double-check Head Fake = NO.\n"
             "If Head Fake = YES (amber row), DO NOT ENTER regardless of the signal."),
            ("Step 6", "Check CMF and MFI values",
             "CMF > +0.10 AND MFI > 80 = maximum conviction → consider larger position.\n"
             "These are shown in bright green if the values are strong."),
            ("Step 7", "Check Squeeze Days",
             "20+ days = spring has been coiling a long time = potentially big move.\n"
             "Green = very long squeeze. Yellow = moderate squeeze."),
            ("Step 8", "Optionally check Fundamentals",
             "If Fund. Score column shows a value:\n"
             "65+ (green) = strong business backing the trade.\n"
             "Below 40 (red) = trade may work short-term but reduce position size."),
        ]),
        ("UNDERSTANDING THE SQUEEZE SCAN (WAIT SIGNALS)", "1A5276", [
            ("Concept", "What is a WAIT stock?",
             "A WAIT stock is in active squeeze (spring coiled) but the breakout hasn't happened yet.\n"
             "These are your PREPARATION LIST — watch them daily."),
            ("Priority 1", "Sort by Squeeze Days (descending)",
             "The longer a stock has been in squeeze, the more energy stored.\n"
             "20+ day squeezes are the highest priority to watch."),
            ("Priority 2", "Look at Direction Lean",
             "BULLISH lean during squeeze = upside breakout likely when it triggers.\n"
             "BEARISH lean = downside breakdown likely. Avoid BEARISH lean stocks for longs."),
            ("Priority 3", "Check Phase column",
             "Phase 2 (DIRECTION CLUES) stocks are more advanced in the squeeze.\n"
             "They're showing institutional positioning already — closer to breakout."),
            ("Priority 4", "Check CMF during squeeze",
             "CMF positive during squeeze = institutions quietly accumulating.\n"
             "When CMF is positive AND squeeze days > 10 = high-quality setup forming."),
            ("Daily Action", "What to do with WAIT stocks",
             "Run the scanner every evening. When a WAIT stock becomes a BUY:\n"
             "1. Check all 5 conditions are GREEN.\n"
             "2. Verify Head Fake = NO.\n"
             "3. Enter the next morning at 9:15 AM NSE open."),
        ]),
        ("UNDERSTANDING THE SELL/EXIT SCAN", "641E16", [
            ("Priority 1", "Exit Signal 1 — SAR Flip (most urgent)",
             "Exit: SAR Flip = YES → EXIT IMMEDIATELY at next morning's open.\n"
             "This is the PRIMARY exit signal. No debate, no waiting for recovery.\n"
             "The trend has officially reversed."),
            ("Priority 2", "Exit Signal 2 — Lower Band Tag",
             "Exit: LowBand = YES → The full move is complete. Take your profits.\n"
             "Price has traveled from upper band to lower band = textbook profit target."),
            ("Priority 3", "Exit Signal 3 — Double Negative",
             "Exit: DblNeg = YES → Early warning. CMF negative + MFI below 50.\n"
             "Smart money exiting + fuel exhausted. Exit to protect profits before SAR flips."),
            ("Key Rule", "One exit signal is enough",
             "You do NOT need all three exit signals to fire.\n"
             "ONE exit signal = exit the next morning. No exceptions.\n"
             "Discipline is more important than trying to squeeze extra profits."),
        ]),
        ("USING FUNDAMENTALS TO SIZE YOUR POSITION", "4A235A", [
            ("Rule 1", "Technicals = WHEN. Fundamentals = HOW MUCH.",
             "The 5 buy conditions tell you WHEN to enter.\n"
             "The fundamental score tells you HOW MUCH capital to risk."),
            ("Rule 2", "High score (65+) = Full position",
             "If Fund. Score >= 65 (green) AND BUY signal → enter with your full planned capital.\n"
             "Strong business + technical setup = highest conviction trade."),
            ("Rule 3", "Medium score (40-64) = Half position",
             "Average fundamentals = the trade may still work technically,\n"
             "but the underlying business is not exceptional. Reduce risk to 50%."),
            ("Rule 4", "Low score (<40) = Smaller or skip",
             "Weak fundamentals = higher risk. Even if technical signal fires,\n"
             "the trade is less reliable. Use 25% position or skip entirely."),
            ("Rule 5", "N/A fundamentals",
             "If Fund. Score = N/A, you chose not to fetch fundamentals during scan.\n"
             "Use Option 1 (single stock analysis) for the specific stock to see fundamentals.\n"
             "Or: run the scan again and say YES to 'Fetch fundamental scores?'"),
        ]),
    ]

    row = 2
    for section_title, section_color, steps in sections:
        ws.merge_cells(f"A{row}:C{row}")
        sec = ws[f"A{row}"]
        sec.value     = section_title
        sec.fill      = _fill(section_color)
        sec.font      = _font(bold=True, color=WHITE, size=11)
        sec.alignment = _align("center", "center")
        ws.row_dimensions[row].height = 20
        row += 1

        for step, title, body in steps:
            ws.cell(row=row, column=1, value=step).fill   = _fill("EAF4FC")
            ws.cell(row=row, column=1).font  = _font(bold=True, size=9, color="1A5276")
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = _align("center", "top")
            ws.cell(row=row, column=2, value=title).fill  = _fill("EAF4FC")
            ws.cell(row=row, column=2).font  = _font(bold=True, size=9)
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = _align("left", "top")
            ws.cell(row=row, column=3, value=body).fill   = _fill("FDFEFE")
            ws.cell(row=row, column=3).font  = _font(size=9)
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = _align("left", "top", wrap=True)
            lines = body.count("\n") + 1
            ws.row_dimensions[row].height = max(15, min(lines * 13, 80))
            row += 1

        row += 1


# ══════════════════════════════════════════════════════════════════
#  MAIN EXPORT FUNCTION
# ══════════════════════════════════════════════════════════════════

def export_to_excel(
    results: list[tuple[SignalResult, Optional[FundamentalData]]],
    mode: str = "ALL",
    output_dir: str = ".",
    single_ticker: str = "",
) -> str:
    """
    Export scan/analysis results to a colour-coded, annotated Excel workbook.

    Parameters
    ----------
    results     : list of (SignalResult, FundamentalData | None) tuples
    mode        : "BUY" | "SELL" | "SQUEEZE" | "ALL" | "SINGLE"
    output_dir  : directory to save the file (default: current dir)
    single_ticker : ticker name if exporting single-stock analysis

    Returns
    -------
    str : absolute path to the saved Excel file
    """
    scan_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_tag  = datetime.now().strftime("%Y%m%d_%H%M")

    # Build filename
    if single_ticker:
        filename = f"BB_Squeeze_{single_ticker.replace('.NS', '')}_{date_tag}.xlsx"
    else:
        filename = f"BB_Squeeze_{mode}_Scan_{date_tag}.xlsx"

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ── Sheet 1: Results ─────────────────────────────────────────
    ws_results = wb.active
    ws_results.title = "RESULTS"
    ws_results.sheet_view.showGridLines = False
    ws_results.page_setup.orientation   = "landscape"
    ws_results.page_setup.fitToPage     = True
    ws_results.page_setup.fitToWidth    = 1
    _write_results_sheet(ws_results, results, mode, scan_date)

    # ── Sheet 2: Legend ───────────────────────────────────────────
    ws_legend = wb.create_sheet("LEGEND")
    ws_legend.sheet_view.showGridLines = False
    _write_legend_sheet(ws_legend)

    # ── Sheet 3: Phase Guide ───────────────────────────────────────
    ws_phase = wb.create_sheet("PHASE GUIDE")
    ws_phase.sheet_view.showGridLines = False
    _write_phase_sheet(ws_phase)

    # ── Sheet 4: How to Read ───────────────────────────────────────
    ws_how = wb.create_sheet("HOW TO READ")
    ws_how.sheet_view.showGridLines  = False
    _write_how_to_sheet(ws_how)

    # ── Tab colours ───────────────────────────────────────────────
    ws_results.sheet_properties.tabColor = "00B050"
    ws_legend.sheet_properties.tabColor  = "2471A3"
    ws_phase.sheet_properties.tabColor   = "CB4335"
    ws_how.sheet_properties.tabColor     = "4A235A"

    wb.save(filepath)
    return filepath
